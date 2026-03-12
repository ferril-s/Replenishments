"""FBA Replenishment Automation Script.

Reads BTS Calcs and Available Qty Whse workbooks, computes case-vs-box pack-unit
quantities, and produces two output files:

  1. Amazon FBA Manifest  (Send-to-Amazon workflow upload)
  2. Warehouse Replenishment Sheet  (packing sheet for the warehouse)

Dependencies: pandas, openpyxl
Usage:        Place source files in the working directory and run:
              python replenish.py
"""

import math
import os
import re
import subprocess
import sys
from datetime import date

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Border, Side

# ---------------------------------------------------------------------------
# Configuration — edit these if your folder layout differs
# ---------------------------------------------------------------------------

CONFIG = {
    "input_dir": ".",
    "output_dir": ".",
    "manifest_output": "FBA_Manifest_{date}.xlsx",
    "replenish_output": "WH_Replenishment_{date}.xlsx",
    "sample_size": None,  # None = process all rows where inv > 0
    "replenish_type": "FBA",  # "FBA" or "WFS" — controls which instruction column to populate
    "run_recalc": True,  # Call recalc.py after saving WH to populate cached formula values for FBA
}

# Items whose Case Qty must be forced to a specific value,
# regardless of what the Available Qty / Item List spreadsheet contains.
CASE_QTY_OVERRIDES = {
    8004: 5,
    8005: 5,
    8006: 5,
}

# Alternating row colours for readability
FILL_LIGHT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _num(value, default=0.0):
    """Safely coerce *value* to float, returning *default* for NaN / None / non-numeric."""
    if value is None:
        return default
    try:
        f = float(value)
        return default if math.isnan(f) else f
    except (TypeError, ValueError):
        return default


def lookup_pack_qty_from_data2(sku_no_prefix, data1_df, data2_df):
    """Look up Case Qty and Box Qty from Data2 using the same chain as WH formulas.

    WH chain: VLOOKUP(SKU, Data1!A:B, 2, 0) -> variant Item# -> VLOOKUP(Item#, Data2!A:C/D)

    Returns dict with keys 'case_qty', 'box_qty' (float or 0), and 'item_key'.
    """
    sku_stripped = str(sku_no_prefix).strip()
    d1_match = data1_df[data1_df.iloc[:, 0].astype(str).str.strip() == sku_stripped]
    if d1_match.empty:
        return {"case_qty": 0, "box_qty": 0, "item_key": None}

    item_key = d1_match.iloc[0].iloc[1]
    d2_col_a = data2_df.iloc[:, 0]
    d2_match = data2_df[d2_col_a.astype(str).str.strip() == str(item_key).strip()]

    if d2_match.empty:
        try:
            num_key = int(float(str(item_key)))
            d2_match = data2_df[pd.to_numeric(d2_col_a, errors="coerce") == num_key]
        except (ValueError, TypeError):
            pass

    if d2_match.empty:
        return {"case_qty": 0, "box_qty": 0, "item_key": item_key}

    case_qty = _num(d2_match.iloc[0].iloc[2], 0)
    box_qty = _num(d2_match.iloc[0].iloc[3], 0) if len(d2_match.columns) > 3 else 0
    return {"case_qty": case_qty, "box_qty": box_qty, "item_key": item_key}


def is_simple_ea_bundle(r):
    """True if bundle has all -EA components (simple 1:1 ratio)."""
    components = r.get("component_details", [])
    if not components:
        return False
    return all(
        re.match(r"^\d+-EA$", c["raw"], re.IGNORECASE) for c in components
    )


# ---------------------------------------------------------------------------
# Output path helper
# ---------------------------------------------------------------------------


def unique_path(path):
    """Return *path* if it doesn't exist; otherwise append _v2, _v3, … until free."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    version = 2
    while True:
        candidate = f"{base}_v{version}{ext}"
        if not os.path.exists(candidate):
            return candidate
        version += 1


# ---------------------------------------------------------------------------
# File scanning
# ---------------------------------------------------------------------------


def find_file(keywords, directory=None):
    """Locate a single .xlsx/.xls file whose name contains *all* given keywords.

    Parameters
    ----------
    keywords : str or list[str]
        Substrings that must ALL appear in the filename (case-insensitive).
        Underscores in a keyword also match the equivalent space.
    directory : str, optional
        Folder to scan.  Defaults to ``CONFIG["input_dir"]``.

    Returns
    -------
    str
        Full path to the matched file.

    Raises
    ------
    FileNotFoundError
        When no file matches the keywords.
    """
    if isinstance(keywords, str):
        keywords = [keywords]
    directory = directory or CONFIG["input_dir"]

    def _matches(filename):
        fn = filename.lower()
        for kw in keywords:
            lo = kw.lower()
            alt = kw.replace("_", " ").lower()
            if lo not in fn and alt not in fn:
                return False
        return True

    candidates = [
        f
        for f in os.listdir(directory)
        if f.lower().endswith((".xlsx", ".xls"))
        and not f.startswith("~$")
        and _matches(f)
    ]

    if not candidates:
        raise FileNotFoundError(
            f"No Excel file matching {keywords} found in "
            f"{os.path.abspath(directory)}"
        )

    if len(candidates) == 1:
        return os.path.join(directory, candidates[0])

    print(f"\nMultiple files match {keywords}:")
    for i, name in enumerate(candidates, 1):
        print(f"  {i}. {name}")
    while True:
        try:
            choice = int(input("Enter number: ")) - 1
            if 0 <= choice < len(candidates):
                return os.path.join(directory, candidates[choice])
        except ValueError:
            pass
        print("Invalid choice — try again.")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


# BTS Calcs format variants: (sheet_name, inventory_column_name)
_BTS_FORMAT_WORKING_SHEET = ("Working Sheet", "Inv to Send from Warehouse")
_BTS_FORMAT_INVENTORY_TO_SEND = ("Inventory to Send", "Inventory to Send")


def _load_bts_calcs(bts_path):
    """Load BTS Calcs and detect which sheet/column holds inventory-to-send.

    Supports two workbook formats:
      - Sheet "Working Sheet", column "Inv to Send from Warehouse"
      - Sheet "Inventory to Send", column "Inventory to Send"

    Returns (bts_df, inv_column_name).
    """
    with pd.ExcelFile(bts_path) as xl:
        sheet_names = xl.sheet_names
    for sheet_name, inv_col in (_BTS_FORMAT_WORKING_SHEET, _BTS_FORMAT_INVENTORY_TO_SEND):
        if sheet_name not in sheet_names:
            continue
        df = pd.read_excel(bts_path, sheet_name=sheet_name)
        if inv_col in df.columns:
            return df, inv_col
    available = ", ".join(sheet_names)
    raise KeyError(
        f"BTS Calcs has no supported sheet+column. Tried: "
        f'"{_BTS_FORMAT_WORKING_SHEET[0]}" + "{_BTS_FORMAT_WORKING_SHEET[1]}", '
        f'"{_BTS_FORMAT_INVENTORY_TO_SEND[0]}" + "{_BTS_FORMAT_INVENTORY_TO_SEND[1]}". '
        f"Available sheets: {available}. Check column names on the intended sheet."
    )


def load_sources():
    """Read every source workbook and return DataFrames + openpyxl workbooks.

    Returns a dict with keys:
        bts_df, bts_inv_column, item_df, data1_df, data2_df, instruction_df,
        bundles_df, replenish_wb, manifest_wb
    """
    bts_path = find_file("BTS_Calcs")
    item_path = find_file(["Available", "Qty", "Price Levels"])
    replenish_tpl_path = find_file(["Replenishment", "FBA"])
    manifest_tpl_path = find_file("ManifestFileUpload")

    print(f"  BTS Calcs     : {os.path.basename(bts_path)}")
    print(f"  Available Qty : {os.path.basename(item_path)}")
    print(f"  Replenish Tpl : {os.path.basename(replenish_tpl_path)}")
    print(f"  Manifest Tpl  : {os.path.basename(manifest_tpl_path)}")

    # Support two BTS Calcs formats: Working Sheet + "Inv to Send from Warehouse"
    # or "Inventory to Send" sheet + "Inventory to Send" column
    bts_df, bts_inv_column = _load_bts_calcs(bts_path)
    print(f"  BTS sheet/col : {bts_inv_column}")

    item_df = pd.read_excel(item_path, sheet_name=0)

    replenish_wb = load_workbook(replenish_tpl_path)
    data1_df = pd.read_excel(replenish_tpl_path, sheet_name="Data1")
    data2_df = pd.read_excel(replenish_tpl_path, sheet_name="Data2")
    instruction_df = pd.read_excel(replenish_tpl_path, sheet_name="Instruction")
    bundles_df = pd.read_excel(replenish_tpl_path, sheet_name="Bundles")

    manifest_wb = load_workbook(manifest_tpl_path)

    return {
        "bts_df": bts_df,
        "bts_inv_column": bts_inv_column,
        "item_df": item_df,
        "data1_df": data1_df,
        "data2_df": data2_df,
        "instruction_df": instruction_df,
        "bundles_df": bundles_df,
        "replenish_wb": replenish_wb,
        "manifest_wb": manifest_wb,
    }


# ---------------------------------------------------------------------------
# Item-number extraction
# ---------------------------------------------------------------------------


def extract_item_number(sku):
    """Return the item number from an SKU, preserving optional -A variant suffix.

    Examples::

        FBA_942-A-CASE     -> "942-A"
        FBA_534-EA         -> 534
        FBA_8006-CASE      -> 8006
        FBA_2227-EACH-UPC  -> 2227
    """
    if pd.isna(sku) or sku is None:
        return None
    clean = str(sku).replace("FBA_", "").strip()
    # Match digits plus optional -X variant (single letter before next hyphen or end)
    m = re.match(r"^(\d+)(-[A-Za-z])(?=-|$)", clean, re.IGNORECASE)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    m = re.match(r"^(\d+)", clean)
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# UoM parsing
# ---------------------------------------------------------------------------


def parse_uom(uom_str):
    """Convert a UoM string to the number of pieces per sellable unit.

    ``SET6`` -> 6,  ``EACH`` -> 1,  ``PACK2`` -> 2.
    Returns 1 when the string contains no digits.
    """
    if not uom_str or pd.isna(uom_str):
        return 1
    digits = re.search(r"\d+", str(uom_str))
    return int(digits.group()) if digits else 1


# ---------------------------------------------------------------------------
# Bundle detection & parsing
# ---------------------------------------------------------------------------


def is_bundle(sku):
    """True if *sku* contains ``+``, indicating a multi-item bundle."""
    return "+" in str(sku)


def round_to_unit_multiple(actual, unit_qty, sku):
    """Round Actual Qty to nearest multiple of unit_qty (CASE or BOX); flag if gap > 20%.

    unit_qty = CASE qty for Pack Unit=CASE, BOX qty for Pack Unit=BOX.
    Ensures we never ship partial packs (e.g. 12 when min BOX is 24).
    """
    if not unit_qty or unit_qty == 0:
        return actual, None
    remainder = actual % unit_qty
    if remainder == 0:
        return actual, None
    lower = actual - remainder
    upper = lower + unit_qty
    rounded = lower if (actual - lower) <= (upper - actual) else upper
    gap_pct = abs(rounded - actual) / actual if actual > 0 else 0
    flag = None
    if gap_pct > 0.20:
        flag = (
            f"⚠️ SKU {sku}: Actual qty {actual} rounded to {rounded} "
            f"(gap: {gap_pct:.0%}) — please recheck recommended qty"
        )
    return rounded, flag


def round_to_case_multiple(recommended, units_per_box, sku):
    """Round recommended qty DOWN to nearest multiple of units_per_box; exception: round up to 1 if floor=0."""
    if units_per_box is None or units_per_box == 0:
        return recommended, None
    lower = math.floor(recommended / units_per_box) * units_per_box
    upper = math.ceil(recommended / units_per_box) * units_per_box
    # Prefer round DOWN; exception: if floor gives 0 but recommended > 0, round up to 1 pack
    rounded = lower if lower > 0 or recommended <= 0 else upper
    gap_pct = abs(rounded - recommended) / recommended if recommended > 0 else 0
    flag = None
    if gap_pct > 0.20:
        flag = (
            f"⚠️ SKU {sku}: Recommended qty {recommended} rounded to {rounded} "
            f"(gap: {gap_pct:.0%}) — please recheck"
        )
    return rounded, flag


def resolve_bundle_case_qty(sku, uom, item_list_df, pack_unit):
    """Resolve Units per box for bundle SKUs from ItemList.

    Bundle SKUs have no CASE/BOX in For WH — look up each component's
    Case Qty or Box Qty in ItemList.

    For equal-ratio bundles (e.g. 534-EA+535-EA): floor(min(qty_list) / uom).
    For weighted-ratio bundles (e.g. 803-SET4+804-SET6): extract ratio from
    -SETn suffix, scale box counts by ratio, total_units = sum(box_count_i * box_qty_i),
    units_per_box = floor(total_units / uom).

    Parameters
    ----------
    sku : str
        Bundle SKU (e.g. 534-EA+535-EA or 803-SET4+804-SET6).
    uom : int
        Pieces per sellable set.
    item_list_df : pandas.DataFrame
        ItemList with columns "Item No.", "Case Qty", "Box Qty".
    pack_unit : str
        "CASE" or "BOX" — which quantity column to use.

    Returns
    -------
    int or None
        Units per box, or None if not a bundle or any component not in ItemList.
    """
    if "+" not in str(sku):
        return None
    parts = str(sku).replace("FBA_", "").split("+")
    qty_col = "Case Qty" if pack_unit == "CASE" else "Box Qty"
    items = []  # list of (item_no, qty, ratio)
    for part in parts:
        part = part.strip()
        item_no = extract_item_number("FBA_" + part if not part.startswith("FBA_") else part)
        if item_no is None:
            continue
        ratio_m = re.search(r"-SET(\d+)", part, re.IGNORECASE)
        ratio = int(ratio_m.group(1)) if ratio_m else 1
        row = item_list_df[item_list_df["Item No."].astype(str).str.strip() == str(item_no).strip()]
        if row.empty:
            print(f"  ⚠️ Bundle SKU {sku}: component item {item_no} not found in ItemList — "
                  "Units per box could not be calculated. Please add item to ItemList.")
            return None
        qty = _num(row.iloc[0].get(qty_col), 0)
        if qty <= 0:
            continue
        items.append((item_no, qty, ratio))
    if not items:
        return None
    ratios = [r for (_, _, r) in items]
    if any(r != 1 for r in ratios):
        g = math.gcd(ratios[0], ratios[1]) if len(ratios) >= 2 else ratios[0]
        for r in ratios[2:]:
            g = math.gcd(g, r)
        box_counts = [r // g for r in ratios]
        total_units = sum(bc * qty for (_, qty, _), bc in zip(items, box_counts))
    else:
        total_units = min(qty for (_, qty, _) in items)
    if uom <= 0:
        return int(total_units)
    return math.floor(total_units / uom)


def parse_bundle_components(sku):
    """Parse a bundle SKU into its component parts.

    Parameters
    ----------
    sku : str
        Full SKU such as ``FBA_534-EA+535-EA`` or ``534-EA+535-EA``.

    Returns
    -------
    list[dict]
        One dict per component with keys ``raw``, ``item_no``, ``ea_per_set``.

    Example::

        parse_bundle_components("FBA_534-EA+535-EA")
        # [{"raw": "534-EA", "item_no": 534, "ea_per_set": 1},
        #  {"raw": "535-EA", "item_no": 535, "ea_per_set": 1}]
    """
    clean = str(sku).replace("FBA_", "")
    parts = clean.split("+")
    components = []
    for part in parts:
        item_no = extract_item_number(part)
        uom_suffix = extract_suffix("FBA_" + part)
        ea = parse_uom(uom_suffix)
        components.append({"raw": part, "item_no": item_no, "ea_per_set": ea})
    return components


# ---------------------------------------------------------------------------
# Pack-unit selection
# ---------------------------------------------------------------------------


def _build_case_only_result(inv_to_send, case_qty, case_dims):
    """Build result for CASE suffix (non-8004/5/6): no rounding, ship exactly inv_to_send cases.

    Each case is 1 sellable unit. WH UoM = case_qty. Manifest: Units per box = 1, Num boxes = adj_qty.
    """
    adj_sets = max(1, int(_num(inv_to_send)))
    num_packs = adj_sets
    cq = int(_num(case_qty, 1))
    total_ea = adj_sets * cq
    dims = case_dims or {}
    return {
        "pack_type": "CASE",
        "pack_qty": cq,
        "sets_per_pack": 1,
        "adj_sets": adj_sets,
        "total_ea": total_ea,
        "num_packs": num_packs,
        "length": math.floor(_num(dims.get("length"))),
        "width": math.floor(_num(dims.get("width"))),
        "height": math.floor(_num(dims.get("height"))),
        "weight": round(_num(dims.get("weight")), 2),
        "wh_uom": cq,
        "manifest_units_per_box": 1,
        "manifest_num_boxes": adj_sets,
        "remainder_flag": None,
    }


def _build_box_suffix_result(inv_to_send, case_qty, box_qty, case_dims):
    """Build result for BOX suffix: round to boxes_per_case, always ship full cases.

    Sellable unit = 1 box. WH UoM = box_qty. Manifest: Units per box = case_qty/box_qty,
    Num boxes = physical cases shipped.
    """
    inv_to_send = _num(inv_to_send)
    cq = int(_num(case_qty, 1))
    bq = int(_num(box_qty, 1))
    if bq <= 0:
        bq = 1
    boxes_per_case = cq // bq
    if boxes_per_case <= 0:
        boxes_per_case = 1

    # Round inv_to_send to nearest multiple of boxes_per_case
    lo = (math.floor(inv_to_send / boxes_per_case)) * boxes_per_case
    hi = lo + boxes_per_case
    if abs(hi - inv_to_send) <= abs(inv_to_send - lo):
        adj_sets = int(hi)
    else:
        adj_sets = int(lo)
    adj_sets = max(adj_sets, boxes_per_case)  # minimum 1 full case

    num_packs = adj_sets // boxes_per_case  # physical cases
    total_ea = adj_sets * bq  # total items (EA)
    dims = case_dims or {}
    return {
        "pack_type": "CASE",
        "pack_qty": cq,
        "sets_per_pack": boxes_per_case,
        "adj_sets": adj_sets,
        "total_ea": total_ea,
        "num_packs": num_packs,
        "length": math.floor(_num(dims.get("length"))),
        "width": math.floor(_num(dims.get("width"))),
        "height": math.floor(_num(dims.get("height"))),
        "weight": round(_num(dims.get("weight")), 2),
        "wh_uom": bq,
        "manifest_units_per_box": boxes_per_case,
        "manifest_num_boxes": num_packs,
        "remainder_flag": None,
    }


def select_pack_unit(inv_to_send, case_qty, box_qty, uom, case_dims, box_dims):
    """Decide CASE vs BOX and compute adjusted quantities and dimensions.

    Parameters
    ----------
    inv_to_send : float
        Number of **sellable sets** to ship (from BTS Calcs).
        NOT individual EA — multiply by *uom* to get EA.
    case_qty : float
        Individual pieces (EA) per case, from Item List.
    box_qty : float or None
        Individual pieces (EA) per box, from Item List.
    uom : int
        Pieces per sellable set (e.g. SET24 → 24, EACH → 1).
    case_dims, box_dims : dict or None
        Length / width / height / weight for the pack unit.

    Decision rules (all comparisons in EA)
    ---------------------------------------
    total_ea_raw = inv_to_send * uom
    threshold    = 0.8 * case_qty

    * total_ea_raw >= threshold             -> CASE
    * total_ea_raw <  threshold AND box_qty -> BOX  (two-pass: if BOX-rounded
      total_ea >= threshold, switch to CASE instead)
    * total_ea_raw <  threshold AND no box  -> CASE  (minimum 1 full case)

    If pack_qty % uom != 0 (partial-pack): skip LCM rounding, keep
    adj_sets = inv_to_send, and use math.ceil for num_packs.  The
    warehouse pulls enough packs and sets aside leftover pieces.
    """
    inv_to_send = _num(inv_to_send)          # sellable sets
    case_qty = _num(case_qty, default=1.0)   # EA per case
    if case_qty == 0:
        case_qty = 1.0
    uom = max(1, int(_num(uom, default=1)))  # EA per set
    has_box = _num(box_qty) > 0

    total_ea_raw = inv_to_send * uom         # convert sets → EA
    threshold = 0.8 * case_qty

    if total_ea_raw >= threshold or not has_box:
        # --- CASE path ---
        cq = int(case_qty)
        sets_per_pack = math.floor(cq / uom) if uom > 0 else cq
        if sets_per_pack <= 0:
            sets_per_pack = 1
        remainder_flag = f"⚠️ Remainder: {cq % uom} unit(s) per case will not form a complete set" if (cq % uom) > 0 else None

        if cq % uom != 0:
            # Partial-pack: pack_qty and UoM are misaligned — keep original
            # quantity and pull enough cases, setting aside leftovers.
            num_packs = max(1, math.ceil(inv_to_send / sets_per_pack))
            adj_sets = sets_per_pack * num_packs
            total_ea = cq * num_packs
        else:
            # Round DOWN to nearest multiple of sets_per_pack; exception: round up to 1 case if floor=0
            case_lo = math.floor(inv_to_send / sets_per_pack)
            if case_lo == 0 and inv_to_send > 0:
                num_packs = 1
            else:
                num_packs = max(0, int(case_lo))
            adj_sets = sets_per_pack * num_packs
            total_ea = cq * num_packs

        dims = case_dims
        pack_type = "CASE"
        pack_qty = case_qty
    else:
        # --- BOX path ---
        bq = int(float(box_qty))
        sets_per_pack = math.floor(bq / uom) if uom > 0 else bq
        if sets_per_pack <= 0:
            sets_per_pack = 1
        remainder_flag = f"⚠️ Remainder: {bq % uom} unit(s) per box will not form a complete set" if (bq % uom) > 0 else None

        if bq % uom != 0:
            # Partial-pack for BOX
            num_packs = max(1, math.ceil(inv_to_send / sets_per_pack))
            adj_sets = sets_per_pack * num_packs
            total_ea = bq * num_packs
        else:
            # Round DOWN to nearest multiple; exception: round up to 1 pack if floor=0
            case_lo = math.floor(inv_to_send / sets_per_pack)
            if case_lo == 0 and inv_to_send > 0:
                num_packs = 1
            else:
                num_packs = max(0, int(case_lo))
            adj_sets = sets_per_pack * num_packs
            total_ea = bq * num_packs

        # Two-pass: if the BOX-rounded quantity now meets the CASE threshold,
        # switch to CASE — shipping one case is more efficient than many boxes.
        if total_ea >= threshold:
            cq = int(case_qty)
            sets_per_pack = math.floor(cq / uom) if uom > 0 else cq
            if sets_per_pack <= 0:
                sets_per_pack = 1
            remainder_flag = f"⚠️ Remainder: {cq % uom} unit(s) per case will not form a complete set" if (cq % uom) > 0 else None
            if cq % uom != 0:
                num_packs = max(1, math.ceil(inv_to_send / sets_per_pack))
                adj_sets = sets_per_pack * num_packs
                total_ea = cq * num_packs
            else:
                # Round DOWN; exception: round up to 1 case if floor=0
                case_lo = math.floor(inv_to_send / sets_per_pack)
                if case_lo == 0 and inv_to_send > 0:
                    num_packs = 1
                else:
                    num_packs = max(0, int(case_lo))
                adj_sets = sets_per_pack * num_packs
                total_ea = cq * num_packs
            dims = case_dims
            pack_type = "CASE"
            pack_qty = case_qty
        else:
            num_packs = total_ea // bq if bq > 0 else 1
            dims = box_dims if box_dims else case_dims
            pack_type = "BOX"
            pack_qty = float(box_qty)
    # For EACH/SET/PACK: WH UoM = uom, manifest units/box = sets per pack, num boxes = num_packs
    return {
        "pack_type": pack_type,
        "pack_qty": int(pack_qty),
        "sets_per_pack": sets_per_pack,
        "adj_sets": int(adj_sets),
        "total_ea": total_ea,
        "num_packs": num_packs,
        "length": math.floor(_num(dims.get("length"))),
        "width": math.floor(_num(dims.get("width"))),
        "height": math.floor(_num(dims.get("height"))),
        "weight": round(_num(dims.get("weight")), 2),
        "wh_uom": uom,
        "manifest_units_per_box": sets_per_pack,
        "manifest_num_boxes": num_packs,
        "remainder_flag": remainder_flag,
    }


# ---------------------------------------------------------------------------
# Bundle: Sellable-Set calculation + processing
# ---------------------------------------------------------------------------


def calculate_sellable_set(components, item_df, inv_to_send):
    """Compute the minimum bundle-set step that yields full packs for every component.

    For each component the pack unit (CASE vs BOX) is chosen using the same
    0.8-threshold rule as regular items.  The per-component minimum step is
    ``pack_qty / gcd(pack_qty, ea_per_set)``.  The overall Sellable Set is
    the LCM of all per-component steps.

    Returns
    -------
    sellable_set : int
    comp_info : list[dict]
        Per-component details: item_no, ea_per_set, pack_qty, pack_type, dims, raw.
    """
    comp_info = []
    min_sets_values = []

    for comp in components:
        item_no = comp["item_no"]
        ea_per_set = comp["ea_per_set"]
        item_row = item_df[item_df["Item No."] == item_no]

        case_qty = 1
        box_qty = 0
        case_dims = {"length": 0, "width": 0, "height": 0, "weight": 0}
        box_dims = None

        if not item_row.empty:
            r = item_row.iloc[0]
            case_qty = int(_num(
                CASE_QTY_OVERRIDES.get(item_no, r.get("Case Qty", 1)), 1
            ))
            if case_qty == 0:
                case_qty = 1
            box_qty = _num(r.get("Box Qty", 0))
            case_dims = {
                "length": r.get("Case Length", 0),
                "width": r.get("Case Width", 0),
                "height": r.get("Case Height", 0),
                "weight": r.get("Case Weight", 0),
            }
            if box_qty > 0:
                box_dims = {
                    "length": r.get("Box Length", 0),
                    "width": r.get("Box Width", 0),
                    "height": r.get("Box Height", 0),
                    "weight": r.get("Box Weight", 0),
                }

        comp_ea_raw = inv_to_send * ea_per_set
        threshold = 0.8 * case_qty
        if comp_ea_raw >= threshold or box_qty <= 0:
            pack_qty = case_qty
            pack_type = "CASE"
            dims = case_dims
        else:
            pack_qty = int(box_qty)
            pack_type = "BOX"
            dims = box_dims or case_dims

        min_sets_i = pack_qty // math.gcd(pack_qty, ea_per_set)
        min_sets_values.append(min_sets_i)

        comp_info.append({
            "item_no": item_no,
            "ea_per_set": ea_per_set,
            "pack_qty": pack_qty,
            "pack_type": pack_type,
            "dims": dims,
            "raw": comp["raw"],
        })

    sellable_set = min_sets_values[0]
    for ms in min_sets_values[1:]:
        sellable_set = (sellable_set * ms) // math.gcd(sellable_set, ms)

    return sellable_set, comp_info


def process_bundle(sku, inv_to_send, item_df, bundles_df):
    """Process a bundle SKU and return a result dict matching ``select_pack_unit`` output.

    Steps
    -----
    1. Parse components from *sku*.
    2. Calculate Sellable Set from component pack quantities.
    3. Round *inv_to_send* (bundle sets) to nearest multiple of Sellable Set
       (closest; ties round up).
    4. Compute per-component packs and EA.
    5. Aggregate into a single-row result.
    """
    components = parse_bundle_components(sku)
    inv_to_send = _num(inv_to_send)
    sellable_set, comp_info = calculate_sellable_set(
        components, item_df, inv_to_send
    )
    bundle_uom = sum(ci["ea_per_set"] for ci in comp_info)
    pack_types = {ci["pack_type"] for ci in comp_info}
    overall_pack_type = "CASE" if "CASE" in pack_types else "BOX"
    units_per_box = resolve_bundle_case_qty(sku, bundle_uom, item_df, overall_pack_type)

    if units_per_box and units_per_box > 0:
        adj_sets, rounding_flag = round_to_case_multiple(inv_to_send, units_per_box, sku)
        adj_sets = max(1, int(adj_sets))
    else:
        lo = max(sellable_set, (math.floor(inv_to_send / sellable_set)) * sellable_set)
        hi = lo + sellable_set
        if abs(hi - inv_to_send) <= abs(inv_to_send - lo):
            adj_sets = int(hi)
        else:
            adj_sets = int(lo)
        rounding_flag = None

    total_packs = 0
    total_ea_all = 0
    component_details = []
    for ci in comp_info:
        comp_ea = adj_sets * ci["ea_per_set"]
        comp_packs = comp_ea // ci["pack_qty"]
        total_packs += comp_packs
        total_ea_all += comp_ea
        component_details.append({**ci, "comp_ea": comp_ea, "comp_packs": comp_packs})

    best = max(component_details, key=lambda c: (
        _num(c["dims"].get("length"))
        * _num(c["dims"].get("width"))
        * _num(c["dims"].get("height"))
    ))

    sets_per_pack = adj_sets // total_packs if total_packs > 0 else adj_sets
    if units_per_box is None or units_per_box <= 0:
        units_per_box = sets_per_pack
    manifest_units_per_box = units_per_box if units_per_box is not None else sets_per_pack
    manifest_num_boxes = (
        math.ceil(adj_sets / manifest_units_per_box)
        if manifest_units_per_box and manifest_units_per_box > 0
        else total_packs
    )
    return {
        "pack_type": overall_pack_type,
        "pack_qty": best["pack_qty"],
        "sets_per_pack": sets_per_pack,
        "adj_sets": adj_sets,
        "total_ea": total_ea_all,
        "num_packs": total_packs,
        "manifest_units_per_box": manifest_units_per_box,
        "manifest_num_boxes": manifest_num_boxes,
        "length": math.floor(_num(best["dims"].get("length"))),
        "width": math.floor(_num(best["dims"].get("width"))),
        "height": math.floor(_num(best["dims"].get("height"))),
        "weight": round(_num(best["dims"].get("weight")), 2),
        "is_bundle": True,
        "bundle_uom": bundle_uom,
        "sellable_set": sellable_set,
        "component_details": component_details,
        "rounding_flag": rounding_flag,
    }


# ---------------------------------------------------------------------------
# Instruction fallback
# ---------------------------------------------------------------------------


def extract_suffix(sku):
    """Extract the UoM-like suffix from an FBA SKU.

    Scans hyphen-delimited tokens for a recognised UoM pattern so that
    variant indicators (e.g. the ``A`` in ``17090-A-EACH``) are skipped.

    ``FBA_2227-EACH-UPC`` -> ``EACH``,  ``FBA_582-SET6`` -> ``SET6``,
    ``FBA_17090-A-EACH`` -> ``EACH``,  ``FBA_5077-A-CASE`` -> ``CASE``
    """
    clean = str(sku).replace("FBA_", "")
    stripped = re.sub(r"^\d+", "", clean).lstrip("-")
    if not stripped:
        return ""
    tokens = stripped.split("-")
    _UOM_RE = re.compile(r"^(EACH|EA|SET\d*|PACK\d*|CASE|BOX)$", re.IGNORECASE)
    for token in tokens:
        if _UOM_RE.match(token):
            return token
    return tokens[0]


# Approved default instruction templates (do NOT use 32 yellow-highlighted SKUs).
_DEFAULT_INSTRUCTIONS = {
    "EACH": "Apply outer box/case label accordingly to match item# (Cover all existing barcodes on the box/case)",
    "SET": "• Put {uom} pcs (different color) per poly bag. Apply FNSKU on the poly bags.\n• Put the poly-bagged sets back to CASE (Cover all existing barcodes on the box/case)",
    "PACK": "• Put {uom} pcs into polybag and apply FNSKU labels on each of the poly bags.\n• Insert the poly-bagged sets back to BOX/CASE (Cover all existing barcodes on the box/case)",
    "CASE": '• Apply label "THIS IS A SET" **DO NOT COVER CASE BARCODE**\n• Apply outer case label accordingly to match item#',
    "BOX": "• For inner box: apply FNSKU label and \"THIS IS A SET\" label\n• Cover all visible barcodes on the box/case\n• Place all the inner boxes back into the CASE.\n• Apply outer case label (Cover barcodes on the case)",
}


def _infer_supply(instruction_text):
    """Infer required supply from instruction text.

    Returns ``True`` if the instructions mention poly-bagging (meaning a
    poly bag supply is needed), ``False`` otherwise.
    """
    if not instruction_text:
        return False
    lower = str(instruction_text).lower()
    return "poly bag" in lower or "polybag" in lower


# Mapping: piece count → most-common poly-bag size (from cross-referencing
# existing Instruction sheet entries).  Used as a fallback when no
# same-item match exists.
_POLY_BAG_BY_COUNT = {
    1: "Poly Bag 6 x 9",
    2: "Poly Bag 6 x 9",
    3: "Poly Bag 6 x 9",
    4: "Poly Bag 14.5 x 19",
    5: "Poly Bag 12 x 15.5",
    6: "Poly Bag 12 x 15.5",
    8: "Poly Bag 12 x 15.5",
    10: "Poly Bag 12 x 15.5",
    12: "Poly Bag 9 x 12",
    25: "Poly Bag 12 x 15.5",
}


def _build_supply_lookups(ws_instruction, col_idx):
    """Scan the Instruction sheet and build two lookup dicts for supply inference.

    Returns
    -------
    item_supply : dict
        ``{item_no_str: supply_str}`` — best poly-bag supply for each item
        number (majority vote across all SKUs sharing the same item number).
    count_supply : dict
        ``{int(piece_count): supply_str}`` — most-common poly-bag size for
        each piece count (SET2 → 2, PACK4 → 4, etc.).
    """
    from collections import defaultdict

    item_votes = defaultdict(lambda: defaultdict(int))  # item -> supply -> count
    count_votes = defaultdict(lambda: defaultdict(int))  # pieces -> supply -> count

    for excel_row in range(2, ws_instruction.max_row + 1):
        sku = ws_instruction.cell(row=excel_row, column=1).value
        supply = ws_instruction.cell(row=excel_row, column=10).value
        instr = ws_instruction.cell(row=excel_row, column=col_idx).value
        if not sku or not supply:
            continue
        supply_str = str(supply).strip().replace("Poly bag", "Poly Bag")
        if "Poly Bag" not in supply_str:
            continue

        sku_str = str(sku).strip()
        m_item = re.match(r"^(\d+)", sku_str)
        item_no = m_item.group(1) if m_item else ""
        m_suffix = re.search(r"-(SET|PACK|EACH|EA|CASE|BOX)(\d*)", sku_str, re.IGNORECASE)
        count = int(m_suffix.group(2)) if (m_suffix and m_suffix.group(2)) else 1

        if item_no:
            item_votes[item_no][supply_str] += 1
        count_votes[count][supply_str] += 1

    item_supply = {
        item: max(votes.items(), key=lambda x: x[1])[0]
        for item, votes in item_votes.items()
    }
    count_supply = {
        cnt: max(votes.items(), key=lambda x: x[1])[0]
        for cnt, votes in count_votes.items()
    }

    return item_supply, count_supply


def _resolve_supply(sku, instruction_text, item_supply, count_supply):
    """Determine the correct poly-bag supply string for a SKU.

    Resolution chain (first match wins):
    1. Same item number already has a poly-bag supply in the Instruction sheet
       → use that exact bag size.
    2. Same piece count (SET6 → 6, PACK2 → 2) has a dominant bag size across
       all existing entries → use that.
    3. Built-in ``_POLY_BAG_BY_COUNT`` table → use the default for that count.
    4. Last resort → ``"Poly Bag"`` (generic, no size — flags the warehouse
       to look it up).

    Returns ``"No Supplies"`` if the instruction doesn't involve poly-bagging.
    """
    if not _infer_supply(instruction_text):
        return "No Supplies"

    sku_str = str(sku).strip()

    # 1. Same item number
    m_item = re.match(r"^(\d+)", sku_str)
    item_no = m_item.group(1) if m_item else ""
    if item_no and item_no in item_supply:
        return item_supply[item_no]

    # 2/3. Piece count lookup
    m_suffix = re.search(r"-(SET|PACK|EACH|EA|CASE|BOX)(\d*)", sku_str, re.IGNORECASE)
    count = int(m_suffix.group(2)) if (m_suffix and m_suffix.group(2)) else 1
    if count in count_supply:
        return count_supply[count]
    if count in _POLY_BAG_BY_COUNT:
        return _POLY_BAG_BY_COUNT[count]

    return "Poly Bag"


def get_instruction_fallback(sku, instruction_df=None):
    """Return approved default packing instruction by UoM suffix when SKU has no instruction.

    Uses approved default templates only (does NOT reference yellow-highlighted SKUs).
    """
    if "+" in str(sku):
        return None  # Skip bundles
    suffix = extract_suffix(sku)
    if not suffix:
        return None

    suffix_upper = suffix.upper()
    if suffix_upper.startswith("EACH") or suffix_upper == "EA":
        key = "EACH"
        uom = 1
    elif suffix_upper.startswith("SET"):
        key = "SET"
        uom = parse_uom(suffix)
    elif suffix_upper.startswith("PACK"):
        key = "PACK"
        uom = parse_uom(suffix)
    elif suffix_upper == "CASE":
        key = "CASE"
        uom = 1
    elif suffix_upper == "BOX":
        key = "BOX"
        uom = 1
    else:
        key = "EACH"
        uom = 1

    template = _DEFAULT_INSTRUCTIONS.get(key, _DEFAULT_INSTRUCTIONS["EACH"])
    instr = template.format(uom=uom)
    return {
        "instruction_fba": instr,
        "instruction_wfs": instr,
        "suffix_used": suffix,
    }


# ---------------------------------------------------------------------------
# Instruction sheet updater
# ---------------------------------------------------------------------------

# Instruction sheet column layout:
#   A=SKU  B=ASIN  C=FNSKU  D=Item  E=FBA SKU
#   F=Instruction FBA  G=AWD  H=Instruction WFS  I=GTIN  J=Supplies
INSTR_COL = {"FBA": 6, "WFS": 8}    # which column to write for each type

# Yellow-ish fill colors to skip (do not use instructions from these cells)
_YELLOW_COLORS = {"FFFF00", "FFFF99", "FFFFCC", "FFFACD", "FFF8DC", "FFFC00", "FFEB9C", "FFFFFF00"}


def _is_yellow_fill(cell):
    """Return True if the cell has a yellow or yellow-ish fill."""
    if cell is None or not hasattr(cell, "fill"):
        return False
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return False
    color = None
    if hasattr(fill, "fgColor") and fill.fgColor:
        if hasattr(fill.fgColor, "rgb") and fill.fgColor.rgb:
            color = str(fill.fgColor.rgb).upper().replace("FF", "")
        elif hasattr(fill.fgColor, "indexed") and fill.fgColor.indexed is not None:
            return fill.fgColor.indexed in (6, 13, 43)
    if color is None and hasattr(fill, "start_color") and fill.start_color:
        if hasattr(fill.start_color, "rgb") and fill.start_color.rgb:
            color = str(fill.start_color.rgb).upper()
    if color:
        color_clean = (color[-6:] if len(color) >= 6 else color).upper()
        if color_clean in _YELLOW_COLORS:
            return True
        if len(color_clean) == 6 and color_clean.startswith("FFFF"):
            try:
                b = int(color_clean[4:6], 16)
                return b < 200
            except ValueError:
                pass
    return False


def get_instruction_from_sheet(replenish_wb, sku_no_prefix, replenish_type):
    """Return first non-yellow instruction for SKU, or None if all yellow or not found."""
    ws = replenish_wb["Instruction"]
    col_idx = INSTR_COL.get(replenish_type, 6)
    for excel_row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=excel_row, column=1).value
        if cell_val and str(cell_val).strip() == sku_no_prefix:
            instr_cell = ws.cell(row=excel_row, column=col_idx)
            if _is_yellow_fill(instr_cell):
                continue
            val = instr_cell.value
            if val and val != 0 and not (isinstance(val, float) and math.isnan(val)):
                return str(val)
    return None


def update_instruction_sheet(replenish_wb, rows, instruction_df, replenish_type):
    """Add or update rows in the Instruction sheet for missing/blank instructions.

    * SKU not in sheet → append a new row with the fallback instruction.
    * SKU in sheet but instruction column for *replenish_type* is blank/0 →
      fill in the fallback instruction value.

    Supply resolution uses a three-tier approach:
    1. Same item number already has a poly-bag supply → reuse that size.
    2. Same piece count has a dominant bag size → use that.
    3. Built-in count→size table as a last resort.

    Also backfills all existing rows that have poly-bag instructions but
    missing/``No Supplies`` supply values.
    """
    ws = replenish_wb["Instruction"]
    col_idx = INSTR_COL.get(replenish_type, 6)

    # --- Build smart supply lookups from existing entries ---
    item_supply, count_supply = _build_supply_lookups(ws, col_idx)

    # Build SKU → Excel row map from the existing Instruction sheet
    sku_row_map = {}
    last_instr_row = 1
    for excel_row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=excel_row, column=1).value
        if cell_val:
            sku_row_map[str(cell_val).strip()] = excel_row
            last_instr_row = excel_row

    next_row = last_instr_row + 1
    added = 0
    updated = 0

    for r in rows:
        fb = r.get("instruction_fallback")
        if not fb:
            continue
        instr_text = fb.get("instruction_fba") if replenish_type == "FBA" else fb.get("instruction_wfs")
        if not instr_text or (isinstance(instr_text, float) and math.isnan(instr_text)):
            continue

        sku_no_prefix = r["sku_no_prefix"]
        supply = _resolve_supply(sku_no_prefix, instr_text, item_supply, count_supply)

        if sku_no_prefix in sku_row_map:
            # SKU exists — check if instruction cell is blank/0
            existing_row = sku_row_map[sku_no_prefix]
            existing_val = ws.cell(row=existing_row, column=col_idx).value
            if not existing_val or existing_val == 0:
                ws.cell(row=existing_row, column=col_idx, value=instr_text)
                supplies_val = ws.cell(row=existing_row, column=10).value
                if not supplies_val or supplies_val == 0:
                    ws.cell(row=existing_row, column=10, value=supply)
                updated += 1
        else:
            # SKU missing — add a new row
            ws.cell(row=next_row, column=1, value=sku_no_prefix)          # A  SKU
            ws.cell(row=next_row, column=2, value=r.get("asin", ""))      # B  ASIN
            ws.cell(row=next_row, column=3, value=r.get("fnsku", ""))     # C  FNSKU
            item_no = extract_item_number(r["sku"])
            ws.cell(row=next_row, column=4, value=item_no)                # D  Item
            ws.cell(row=next_row, column=5, value=r["sku"])               # E  FBA SKU
            ws.cell(row=next_row, column=col_idx, value=instr_text)       # F or H
            ws.cell(row=next_row, column=10, value=supply)                # J  Supplies
            sku_row_map[sku_no_prefix] = next_row
            next_row += 1
            added += 1

    if added or updated:
        print(f"  Instruction sheet: {added} new row(s), {updated} updated row(s)")

    # --- Backfill empty supplies for ALL existing rows whose instructions need them ---
    # Rebuild lookups (may have been enriched by the additions above)
    item_supply, count_supply = _build_supply_lookups(ws, col_idx)
    backfilled = 0
    for excel_row in range(2, ws.max_row + 1):
        sku_cell = ws.cell(row=excel_row, column=1).value
        instr_cell = ws.cell(row=excel_row, column=col_idx).value
        supply_cell = ws.cell(row=excel_row, column=10).value
        if not instr_cell or not sku_cell:
            continue
        supply_str = str(supply_cell).strip() if supply_cell else ""
        if supply_str in ("", "0", "No Supplies", "No supplies"):
            needs_poly = _infer_supply(instr_cell)
            if needs_poly:
                resolved = _resolve_supply(
                    str(sku_cell).strip(), instr_cell,
                    item_supply, count_supply,
                )
                ws.cell(row=excel_row, column=10, value=resolved)
                backfilled += 1
    if backfilled:
        print(f"  Instruction sheet: {backfilled} supply value(s) backfilled "
              f"(item-match + count-based sizing)")


# ---------------------------------------------------------------------------
# Data1 / Data2 sheet auto-population
# ---------------------------------------------------------------------------


def _detect_col_type(ws, column, start_row=2, sample=20):
    """Return ``'str'`` if the first non-empty values in *column* are strings, else ``'num'``."""
    for excel_row in range(start_row, min(start_row + sample, ws.max_row + 1)):
        val = ws.cell(row=excel_row, column=column).value
        if val is not None:
            return "str" if isinstance(val, str) else "num"
    return "num"


def update_data1_sheet(replenish_wb, rows):
    """Append rows to the Data1 sheet for any SKUs not already present.

    Data1 layout: A=SKU, B=ITEM (item number), C=UOM (numeric EA per set).
    UoM comes from the result dict (already resolved for CASE/BOX products).
    Bundle SKUs are skipped because they won't match the single-item
    VLOOKUP pattern.

    Column B (ITEM) is written in the same type (int vs str) as the existing
    Data1 rows so that downstream VLOOKUPs into Data2 remain type-consistent.
    """
    ws = replenish_wb["Data1"]

    item_col_type = _detect_col_type(ws, column=2)

    existing_skus: set[str] = set()
    for excel_row in range(2, ws.max_row + 1):
        val = ws.cell(row=excel_row, column=1).value
        if val:
            existing_skus.add(str(val).strip())

    next_row = ws.max_row + 1
    added = 0
    added_skus: list[str] = []

    for r in rows:
        if r.get("is_bundle"):
            continue
        sku_no_prefix = r["sku_no_prefix"]
        if sku_no_prefix in existing_skus:
            continue

        item_no = extract_item_number(r["sku"])
        uom_val = r.get("wh_uom", r.get("uom", 1))

        item_val = str(item_no) if item_col_type == "str" else item_no

        ws.cell(row=next_row, column=1, value=sku_no_prefix)       # A  SKU
        ws.cell(row=next_row, column=2, value=item_val)             # B  ITEM
        ws.cell(row=next_row, column=3, value=uom_val)              # C  UOM (numeric)

        existing_skus.add(sku_no_prefix)
        added_skus.append(sku_no_prefix)
        next_row += 1
        added += 1

    if added:
        print(f"  Data1 sheet: {added} new row(s) added (ITEM col type: {item_col_type})")
        for s in added_skus:
            print(f"    + {s}")


def update_data2_sheet(replenish_wb, rows, item_df):
    """Append rows to the Data2 sheet for any Item numbers not already present,
    and overwrite existing rows whose Case Qty / Box Qty cells contain external
    VLOOKUP formulas or are empty (these would fail when the workbook is
    opened standalone and cause columns N and O in *For WH* to stay blank).

    Data2 layout: A=Item No., B=Item Description, C=Case Qty, D=Box Qty.
    Case/Box quantities come from the Item List (with CASE_QTY_OVERRIDES).
    Item Description is left blank for auto-added rows.

    Column A (Item No.) is ALWAYS written as int so that the plain VLOOKUP
    in 'For WH' columns N and O can reliably match — mixing text and numeric
    item numbers in column A causes VLOOKUP to miss rows silently.
    """
    ws = replenish_wb["Data2"]

    # --- Normalize numeric strings in Data2 col A to int; keep variants like "942-A" as str ---
    normalized_count = 0
    for excel_row in range(2, ws.max_row + 1):
        val = ws.cell(row=excel_row, column=1).value
        if isinstance(val, str) and val.strip():
            try:
                n = int(float(val.strip()))
                ws.cell(row=excel_row, column=1, value=n)
                normalized_count += 1
            except (TypeError, ValueError):
                pass
    if normalized_count:
        print(f"  Data2 sheet: normalized {normalized_count} text item number(s) in col A to int")

    def _item_key(val):
        """Normalize item_no for dict key: int for plain numbers, str for variants like 942-A."""
        if val is None:
            return None
        try:
            return int(float(str(val).strip()))
        except (TypeError, ValueError):
            return str(val).strip()

    # Build map of item_no -> excel_row for all existing rows
    existing_item_row: dict = {}
    for excel_row in range(2, ws.max_row + 1):
        val = ws.cell(row=excel_row, column=1).value
        if val is not None:
            k = _item_key(val)
            if k is not None:
                existing_item_row[k] = excel_row

    all_item_nos = set()
    for r in rows:
        if r.get("is_bundle"):
            for cd in r.get("component_details", []):
                if cd.get("item_no") is not None:
                    all_item_nos.add(cd["item_no"])
        else:
            item_no = extract_item_number(r["sku"])
            if item_no is not None:
                all_item_nos.add(item_no)

    next_row = ws.max_row + 1
    added = 0
    updated = 0

    for item_no in sorted(all_item_nos, key=lambda x: (0, x) if isinstance(x, int) else (1, str(x))):
        case_qty = CASE_QTY_OVERRIDES.get(item_no)
        box_qty = None

        item_lookup_key = _item_key(item_no)
        if case_qty is None:
            match = item_df[item_df["Item No."].astype(str).str.strip() == str(item_no).strip()]
            if not match.empty:
                case_qty = _num(match.iloc[0].get("Case Qty"), 0)
                box_qty = _num(match.iloc[0].get("Box Qty"), 0)
            else:
                case_qty = 0
                box_qty = 0
        else:
            match = item_df[item_df["Item No."].astype(str).str.strip() == str(item_no).strip()]
            if not match.empty:
                box_qty = _num(match.iloc[0].get("Box Qty"), 0)
            else:
                box_qty = 0

        case_val = int(case_qty) if case_qty else ""
        box_val = int(box_qty) if box_qty else ""
        item_val = item_no  # always int — column A is normalized to int above

        if item_lookup_key is not None and item_lookup_key in existing_item_row:
            # Row already exists — overwrite C and D only when the current
            # value is a formula string (e.g. external VLOOKUP) or missing.
            er = existing_item_row[item_lookup_key]
            existing_c = ws.cell(row=er, column=3).value
            existing_d = ws.cell(row=er, column=4).value
            needs_update = False
            if existing_c is None or (isinstance(existing_c, str) and existing_c.startswith("=")):
                ws.cell(row=er, column=3, value=case_val)
                needs_update = True
            if existing_d is None or (isinstance(existing_d, str) and existing_d.startswith("=")):
                ws.cell(row=er, column=4, value=box_val)
                needs_update = True
            if needs_update:
                updated += 1
        else:
            # Item missing — add a new row
            ws.cell(row=next_row, column=1, value=item_val)                          # A
            ws.cell(row=next_row, column=2, value="")                                # B
            ws.cell(row=next_row, column=3, value=case_val)                          # C
            ws.cell(row=next_row, column=4, value=box_val)                           # D
            next_row += 1
            added += 1

    if added or updated:
        print(f"  Data2 sheet: {added} new row(s) added, {updated} row(s) updated")
        new_items = all_item_nos - set(existing_item_row.keys())
        for item_no in sorted(new_items, key=lambda x: (0, x) if isinstance(x, int) else (1, str(x))):
            print(f"    + Item #{item_no}")


# ---------------------------------------------------------------------------
# Worksheet-name helper
# ---------------------------------------------------------------------------


def find_sheet(wb, search_term):
    """Return the first worksheet whose name contains *search_term* (case-insensitive)."""
    for name in wb.sheetnames:
        if search_term.lower() in name.lower():
            return wb[name]
    raise KeyError(
        f"No sheet matching '{search_term}' in workbook.  "
        f"Available sheets: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# Output 1 — Amazon FBA Manifest
# ---------------------------------------------------------------------------


MANIFEST_HEADERS = [
    "Merchant SKU",       # A
    "Quantity",           # B
    "Prep owner",         # C  (values left blank)
    "Labeling owner",     # D  (values left blank)
    "Units per box",      # E
    "Number of boxes",    # F
    "Box length (in)",    # G
    "Box width (in)",     # H
    "Box height (in)",    # I
    "Box weight (lb)",    # J
    "Pack Unit",          # K
]


def validate_manifest_num_boxes(rows):
    """Check that Number of boxes = Sellable Unit / Units per box for each row.

    Logs mismatches to console. For CASE: num_boxes should equal Case to pull.
    For BOX: num_boxes should equal Box to pull.
    """
    mismatches = []
    for r in rows:
        sellable = r.get("adj_sets", 0)
        upb = r.get("manifest_units_per_box") or r.get("sets_per_pack")
        num_boxes = r.get("manifest_num_boxes") or r.get("num_packs")
        if upb is None or upb <= 0:
            continue
        expected = sellable / upb
        if abs((num_boxes or 0) - expected) > 0.01:
            mismatches.append(
                f"SKU {r.get('sku', '?')}: Number of boxes mismatch — "
                f"has {num_boxes}, expected {expected:.2f}"
            )
    if mismatches:
        print("\n  FBA Manifest Number of boxes validation:")
        for msg in mismatches:
            print(f"    ⚠️ {msg}")


def _recalc_and_read_wh(wh_path, num_rows):
    """Run recalc.py on the WH file and read cached formula values.

    Returns a DataFrame with WH sheet data (header row + data rows), or None
    if recalc fails (e.g. Excel not available).
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    recalc_script = os.path.join(script_dir, "recalc.py")
    if not os.path.isfile(recalc_script):
        return None
    try:
        subprocess.run(
            [sys.executable, recalc_script, os.path.abspath(wh_path)],
            check=True,
            capture_output=True,
            cwd=script_dir,
        )
    except (subprocess.CalledProcessError, FileNotFoundError):
        return None

    try:
        # Read with data_only so we get cached calculated values after recalc
        wb = load_workbook(wh_path, data_only=True, read_only=True)
        ws = wb["For WH"]
        # Build list of rows: header + data (skip header when building row list for lookup)
        # WH columns: A=0 SKU, F=5 Sellable, G=6 Box pull, H=7 Case pull, I=8 Sets, S=18 Pack Unit
        data = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=1 + num_rows, values_only=True)):
            data.append(row)
        wb.close()
        return data
    except Exception:
        return None


def write_manifest(manifest_wb, rows, output_path, wh_data=None, data1_df=None, data2_df=None):
    """Create a clean manifest workbook using the Amazon Send-to-Amazon headers.

    Column headers match the official ManifestFileUpload template exactly.
    Prep owner (C) and Labeling owner (D) are left blank per Amazon defaults.
    Pack Unit (K) is appended so the warehouse knows CASE vs BOX.

    If wh_data is provided (list of tuples from recalculated WH sheet), FBA
    Quantity/Units per box/Number of boxes are read from WH cached values
    (cols F, I, G or H) instead of computed from rows.
    """
    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "FBA Manifest"

    # Row 1: headers
    for col, header in enumerate(MANIFEST_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD

    # WH columns: A=0 SKU, F=5 Sellable, G=6 Box pull, H=7 Case pull, I=8 Sets, S=18 Pack Unit
    WH_COL_SELLABLE, WH_COL_BOX, WH_COL_CASE, WH_COL_SETS, WH_COL_PACK = 5, 6, 7, 8, 18

    # Row 2+: data
    for idx, r in enumerate(rows):
        rn = 2 + idx
        pack_type = r.get("pack_type", "")
        # Prefer WH cached values when available
        if wh_data and idx < len(wh_data):
            wh_row = wh_data[idx]
            wh_sellable = wh_row[WH_COL_SELLABLE] if len(wh_row) > WH_COL_SELLABLE else None
            wh_sets = wh_row[WH_COL_SETS] if len(wh_row) > WH_COL_SETS else None
            wh_box = wh_row[WH_COL_BOX] if len(wh_row) > WH_COL_BOX else None
            wh_case = wh_row[WH_COL_CASE] if len(wh_row) > WH_COL_CASE else None
            wh_pack = wh_row[WH_COL_PACK] if len(wh_row) > WH_COL_PACK else pack_type
            quantity = _num(wh_sellable) if wh_sellable is not None and wh_sellable != "" else r["adj_sets"]
            units_per_box = _num(wh_sets) if wh_sets is not None and wh_sets != "" else None
            if units_per_box is None or units_per_box <= 0:
                units_per_box = None
            use_case = (str(wh_pack or "").upper() == "CASE")
            num_boxes = _num(wh_case if use_case else wh_box)
            if num_boxes is None or (isinstance(num_boxes, float) and math.isnan(num_boxes)):
                num_boxes = None
        else:
            quantity = r["adj_sets"]
            units_per_box = None
            num_boxes = None

        # Bundle rows: simple EA bundles get computed values; complex stay blank
        if r.get("is_bundle") and not is_simple_ea_bundle(r):
            quantity = None
            units_per_box = None
            num_boxes = None
        else:
            # Fallback to row computation when WH data unavailable or blank
            # Runs for non-bundles and simple EA bundles
            if units_per_box is None or num_boxes is None:
                if r.get("is_bundle") and r.get("manifest_units_per_box") and r.get("manifest_num_boxes"):
                    units_per_box = r["manifest_units_per_box"]
                    num_boxes = r["manifest_num_boxes"]
                    quantity = r["adj_sets"]
                else:
                    # Use Data2 via Data1 chain (same as WH formulas) when available
                    d2_pack_qty = 0
                    if data1_df is not None and data2_df is not None:
                        d2_vals = lookup_pack_qty_from_data2(
                            r["sku_no_prefix"], data1_df, data2_df
                        )
                        pack_unit = r.get("pack_type", "CASE")
                        if pack_unit == "CASE":
                            d2_pack_qty = d2_vals["case_qty"]
                        else:
                            d2_pack_qty = d2_vals["box_qty"] if d2_vals["box_qty"] > 0 else d2_vals["case_qty"]

                    uom = _num(r.get("uom", r.get("wh_uom", 1)), 1)
                    if uom <= 0:
                        uom = 1

                    if d2_pack_qty > 0:
                        units_per_box = math.floor(d2_pack_qty / uom)
                    else:
                        pack_qty = _num(r.get("pack_qty"), 0)
                        units_per_box = math.floor(pack_qty / uom) if pack_qty and uom else r.get("sets_per_pack", 1)

                    if units_per_box <= 0:
                        units_per_box = 1
                    num_boxes = math.ceil(r["adj_sets"] / units_per_box)
                    quantity = units_per_box * num_boxes  # Sellable = Sets x pull count
            if quantity is None or (isinstance(quantity, float) and math.isnan(quantity)):
                quantity = r["adj_sets"]

        is_bundle_blank = r.get("is_bundle") and not is_simple_ea_bundle(r)
        ws.cell(row=rn, column=1, value=r["sku"])          # A  Merchant SKU
        ws.cell(row=rn, column=2, value="" if is_bundle_blank else int(quantity))  # B  Quantity
        # C  Prep owner — blank
        # D  Labeling owner — blank
        ws.cell(row=rn, column=5, value="" if is_bundle_blank else int(units_per_box))  # E  Units per box
        ws.cell(row=rn, column=6, value="" if is_bundle_blank else int(num_boxes))      # F  Number of boxes
        ws.cell(row=rn, column=7, value="" if is_bundle_blank else r["length"])   # G  Box length
        ws.cell(row=rn, column=8, value="" if is_bundle_blank else r["width"])    # H  Box width
        ws.cell(row=rn, column=9, value="" if is_bundle_blank else r["height"])   # I  Box height
        ws.cell(row=rn, column=10, value="" if is_bundle_blank else r["weight"])  # J  Box weight
        ws.cell(row=rn, column=11, value=r["pack_type"])    # K  Pack Unit

        fill = FILL_LIGHT if idx % 2 == 0 else FILL_WHITE
        for col in range(1, 12):
            ws.cell(row=rn, column=col).fill = fill

    # Auto-fit column widths for readability
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 12):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 12)

    out_wb.save(output_path)
    print(f"  Manifest saved  -> {output_path}")


# ---------------------------------------------------------------------------
# Output 2 — Warehouse Replenishment Sheet
# ---------------------------------------------------------------------------


def write_replenishment(replenish_wb, rows, output_path, removal_log=None, data1_df=None, data2_df=None, item_to_avail_qty=None):
    """Fill the *For WH* sheet with data + formulas and save as a new file.

    Hard-coded columns: A (SKU), B (ASIN), C (FNSKU), P (Rec Qty), R (Pack Unit).
    All other columns are live Excel formulas that calculate when the file is
    opened in Excel.

    If removal_log is provided, adds a 'Removed Items' sheet listing filtered rows.
    """
    ws = replenish_wb["For WH"]

    # Extra column headers
    ws.cell(row=1, column=17, value="Actual Qty Replenished").font = BOLD  # Q
    ws.cell(row=1, column=18, value="Current WH Inventory").font = BOLD    # R
    ws.cell(row=1, column=19, value="Pack Unit").font = BOLD               # S
    ws.cell(row=1, column=20, value="Inventory Flag").font = BOLD          # T

    last_data_row = 1 + len(rows)

    for idx, r in enumerate(rows):
        rn = 2 + idx          # Excel row number
        rn_s = str(rn)        # stringified for formula interpolation

        # --- Hard-coded input columns ---
        ws.cell(row=rn, column=1, value=r["sku_no_prefix"])        # A  SKU
        ws.cell(row=rn, column=2, value=r["asin"])                 # B  ASIN
        ws.cell(row=rn, column=3, value=r["fnsku"])                # C  FNSKU
        ws.cell(row=rn, column=16, value=r["inv_to_send"])        # P  Rec Replenishment Qty
        ws.cell(row=rn, column=17, value="" if (r.get("is_bundle") and not is_simple_ea_bundle(r)) else r["adj_sets"])  # Q  Actual
        ws.cell(row=rn, column=19, value=r["pack_type"])           # S  Pack Unit

        if r.get("is_bundle"):
            components = r.get("component_details", [])
            item_nums = ", ".join(str(cd["item_no"]) for cd in components)

            if is_simple_ea_bundle(r):
                # --- Simple EA bundle: write computed values ---
                ws.cell(row=rn, column=4, value=item_nums)                    # D  Item #
                ws.cell(row=rn, column=5, value=r["total_ea"])                # E  Total EA
                ws.cell(row=rn, column=6, value=r["adj_sets"])                # F  Sellable
                if r["pack_type"] == "BOX":
                    ws.cell(row=rn, column=7, value=r.get("num_packs", ""))   # G  Box pull
                    ws.cell(row=rn, column=8, value="")                        # H  Case pull
                else:
                    ws.cell(row=rn, column=7, value="")                        # G  Box pull
                    ws.cell(row=rn, column=8, value=r.get("num_packs", ""))   # H  Case pull
                upb = r.get("manifest_units_per_box", r.get("sets_per_pack", 1))
                ws.cell(row=rn, column=9, value=upb)                          # I  Sets
                ws.cell(row=rn, column=10,
                        value=f'=IF(C{rn_s}="","",IF(B{rn_s}=C{rn_s},"","Labeling"))')  # J Label
                ws.cell(row=rn, column=11, value="")                          # K  Instructions (manual)
                ws.cell(row=rn, column=12, value="")                          # L  Supplies (manual)
                ws.cell(row=rn, column=13, value=r.get("bundle_uom", 1))     # M  UoM (hardcoded)
                # N and O: hardcode from component Data2 (VLOOKUP fails for comma Item#)
                min_case = None
                min_box = None
                for cd in components:
                    d2 = lookup_pack_qty_from_data2(
                        cd["raw"], data1_df, data2_df
                    ) if data1_df is not None else {"case_qty": 0, "box_qty": 0}
                    cq = d2.get("case_qty") or _num(cd.get("pack_qty"), 0)
                    bq = d2.get("box_qty", 0)
                    if cq > 0:
                        min_case = min(min_case, cq) if min_case is not None else cq
                    if bq > 0:
                        min_box = min(min_box, bq) if min_box is not None else bq
                ws.cell(row=rn, column=14, value=int(min_box) if min_box is not None else "")   # N  BOX
                ws.cell(row=rn, column=15, value=int(min_case) if min_case is not None else "")  # O  CASE
            else:
                # --- Complex bundle: leave blank for manual entry ---
                ws.cell(row=rn, column=4, value=item_nums)
                for blank_col in (5, 6, 7, 8, 9, 11, 12):
                    ws.cell(row=rn, column=blank_col, value="")
                ws.cell(row=rn, column=10,
                        value=f'=IF(C{rn_s}="","",IF(B{rn_s}=C{rn_s},"","Labeling"))')
                ws.cell(row=rn, column=13,
                        value=f'=IFERROR(VLOOKUP(A{rn_s},Data1!A:C,3,0),1)')
                ws.cell(row=rn, column=14, value="")  # N  BOX — blank for complex
                ws.cell(row=rn, column=15, value="")  # O  CASE — blank for complex
        else:
            # --- Regular row: E, F, I as Excel formulas (V5); G, H from formulas ---
            ws.cell(row=rn, column=4,                                  # D  Item #
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Data1!A:B,2,0),"")')

            ws.cell(row=rn, column=5,                                  # E  Total EA = M*Q
                    value=f"=M{rn_s}*Q{rn_s}")

            ws.cell(row=rn, column=6,                                  # F  Sellable Unit = I*(H or G)
                    value=f'=I{rn_s}*IF(S{rn_s}="CASE",H{rn_s},G{rn_s})')

            # Mutually exclusive: Pack Unit=CASE → only Case to pull; Pack Unit=BOX → only Box to pull
            ws.cell(row=rn, column=7,                                  # G  Box to pull
                    value=f'=IF(S{rn_s}="CASE","",IFERROR(ROUNDUP(E{rn_s}/N{rn_s},0),""))')

            ws.cell(row=rn, column=8,                                  # H  Case to pull
                    value=f'=IF(S{rn_s}="BOX","",IFERROR(ROUNDUP(E{rn_s}/O{rn_s},0),""))')

            ws.cell(row=rn, column=9,                                  # I  Sets per Box/Case
                    value=f'=IF(S{rn_s}="CASE",FLOOR(O{rn_s}/M{rn_s},1),FLOOR(N{rn_s}/M{rn_s},1))')

            ws.cell(row=rn, column=10,                                 # J  Label
                    value=f'=IF(C{rn_s}="","",IF(B{rn_s}=C{rn_s},"","Labeling"))')

            ws.cell(row=rn, column=11,                                 # K  Packing Instructions
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Instruction!A:F,6,0),"")')

            ws.cell(row=rn, column=12,                                 # L  Supplies
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Instruction!A:J,10,0),"No Supplies")')

            ws.cell(row=rn, column=13,                                 # M  UoM
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Data1!A:C,3,0),1)')

        # N and O: VLOOKUP from Data2; only for regular rows (bundles set N/O in their block)
        if not r.get("is_bundle"):
            ws.cell(row=rn, column=14,                                 # N  BOX
                    value=f'=IFERROR(VLOOKUP(D{rn_s},Data2!$A:$D,4,0),"")')
            ws.cell(row=rn, column=15,                                 # O  CASE
                    value=f'=IFERROR(VLOOKUP(D{rn_s},Data2!$A:$C,3,0),"")')

        # R  Current WH Inventory — VLOOKUP to external warehouse file
        ws.cell(
            row=rn,
            column=18,
            value=f"=IFERROR(VLOOKUP(D{rn_s},'S:\\!E-Commerce\\AMAZON\\Pricing\\price&qty\\[Available Qty Whse 01 + Price Levels.xlsx]Sheet1'!$B:$D,3,0),\"\")",
        )

        # T  Inventory Flag — warns when Total EA exceeds Current WH Inventory
        ws.cell(
            row=rn,
            column=20,
            value=f'=IF(OR(E{rn_s}="",R{rn_s}="",R{rn_s}=0),"",IF(E{rn_s}>R{rn_s},"OVER WH INV",""))',
        )

        # Alternating row fill + thin borders
        fill = FILL_LIGHT if idx % 2 == 0 else FILL_WHITE
        for col in range(1, 21):
            cell = ws.cell(row=rn, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER

    # --- Conditional formatting: highlight R2:R{last_data_row} when inventory < 100 ---
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws.conditional_formatting.add(
        f"R2:R{last_data_row}",
        CellIsRule(operator="lessThan", formula=["100"], fill=red_fill),
    )

    # --- Conditional formatting: highlight T (flag) and E when Total EA > Current WH Inventory ---
    orange_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    orange_font = Font(bold=True, color="FFFFFF")
    ws.conditional_formatting.add(
        f"T2:T{last_data_row}",
        FormulaRule(
            formula=[f'AND(E2<>"",R2<>"",R2>0,E2>R2)'],
            fill=orange_fill,
            font=orange_font,
        ),
    )
    ws.conditional_formatting.add(
        f"E2:E{last_data_row}",
        FormulaRule(
            formula=[f'AND(E2<>"",R2<>"",R2>0,E2>R2)'],
            fill=orange_fill,
        ),
    )

    # --- Remainder highlighting: amber SKU cells where pack_qty % uom != 0 ---
    amber_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    for idx, r in enumerate(rows):
        if r.get("is_bundle") or "+" in str(r.get("sku_no_prefix", "")):
            continue
        pack_qty = _num(r.get("pack_qty"), 0)
        uom = _num(r.get("uom", r.get("wh_uom", 1)), 1)
        if uom <= 0 or pack_qty <= 0:
            continue
        if pack_qty % uom != 0:
            rn = 2 + idx
            ws.cell(row=rn, column=1).fill = amber_fill

    # --- Totals row ---
    tr = last_data_row + 1
    ws.cell(row=tr, column=1, value="TOTALS").font = BOLD
    for col_idx in (5, 6, 7, 8):  # E Total EA, F Sellable, G Box pull, H Case pull (NOT I)
        letter = chr(ord("A") + col_idx - 1)
        cell = ws.cell(row=tr, column=col_idx,
                       value=f"=SUM({letter}2:{letter}{last_data_row})")
        cell.font = BOLD

    # Remove stray formulas from template (e.g. =SUM(G997:H997) in row below TOTALS)
    for clear_row in (tr + 1, tr + 2):
        for clear_col in (7, 8):  # G, H
            c = ws.cell(row=clear_row, column=clear_col)
            if c.value and isinstance(c.value, str) and c.value.startswith("="):
                c.value = None

    # --- Update Table1 reference to cover header + data + totals ---
    try:
        table = ws.tables["Table1"]
        table.ref = f"A1:T{tr}"
    except (KeyError, AttributeError):
        for tname in list(ws.tables):
            ws.tables[tname].ref = f"A1:T{tr}"
            break

    # --- Removed Items sheet ---
    if removal_log:
        if "Removed Items" in replenish_wb.sheetnames:
            del replenish_wb["Removed Items"]
        rm_ws = replenish_wb.create_sheet("Removed Items")
        rm_ws.cell(row=1, column=1, value="SKU").font = BOLD
        rm_ws.cell(row=1, column=2, value="Reason").font = BOLD
        rm_ws.cell(row=1, column=3, value="Value").font = BOLD
        for idx, entry in enumerate(removal_log):
            rn = 2 + idx
            rm_ws.cell(row=rn, column=1, value=str(entry.get("SKU", "")))
            rm_ws.cell(row=rn, column=2, value=str(entry.get("Reason", "")))
            rm_ws.cell(row=rn, column=3, value=entry.get("Value", ""))

    replenish_wb.save(output_path)
    print(f"  Replenishment saved -> {output_path}")


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------


def print_summary(rows, missing_instructions, missing_pack_data):
    """Print a human-readable processing summary to stdout."""
    total_items = len(rows)
    total_ea = sum(
        r["total_ea"] for r in rows
        if not r.get("is_bundle") or is_simple_ea_bundle(r)
    )
    cases = [r for r in rows if r["pack_type"] == "CASE"]
    boxes = [r for r in rows if r["pack_type"] == "BOX"]
    bundles = [r for r in rows if r.get("is_bundle")]

    print(f"\n{'=' * 58}")
    print(f"  Total items processed  : {total_items}")
    print(f"  Total EA               : {total_ea:,}")
    print(f"  CASE decisions         : {len(cases)}")
    print(f"  BOX  decisions         : {len(boxes)}")
    print(f"  Bundles                : {len(bundles)}")

    if bundles:
        print("  Bundle details:")
        for r in bundles:
            comps = ", ".join(
                f"{cd['item_no']}×{cd['comp_packs']}{cd['pack_type'][0]}"
                for cd in r["component_details"]
            )
            print(f"    - {r['sku']}  adj_sets={r['adj_sets']}  "
                  f"sellable_set={r['sellable_set']}  [{comps}]")

    if boxes:
        print("  Items shipped as BOX:")
        for r in boxes:
            if not r.get("is_bundle"):
                print(f"    - {r['sku']}  ({r['total_ea']} EA, "
                      f"{r['num_packs']} box(es))")

    print(f"{'=' * 58}")

    if missing_instructions:
        print("\n  Instruction fallbacks used:")
        for m in missing_instructions:
            fba_val = m.get("instruction_fba", "N/A")
            print(f"    {m['sku']}  -> suffix '{m['suffix_used']}'  "
                  f"-> \"{fba_val}\"")

    if missing_pack_data:
        print("\n  WARNING — no pack-unit data found in Item List for:")
        for sku in missing_pack_data:
            print(f"    - {sku}")

    remainder_warnings = [r for r in rows if r.get("remainder_flag")]
    if remainder_warnings:
        print("\n  Remainder warnings (case/box not a clean multiple of UoM):")
        for r in remainder_warnings:
            print(f"    - {r['sku']}: {r['remainder_flag']}")

    rounding_flags = [r for r in rows if r.get("rounding_flag")]
    if rounding_flags:
        print("\n  Bundle rounding warnings (gap > 20%):")
        for r in rounding_flags:
            print(f"    - {r['rounding_flag']}")

    print()


def print_inventory_flag_report(rows, item_to_avail_qty):
    """Print a terminal report of items where Total EA exceeds Current WH Inventory.

    This catches cases where the replenishment quantity is larger than what the
    warehouse actually has on hand — a shortage that must be resolved before
    shipping.

    Parameters
    ----------
    rows : list[dict]
        Processed row dicts (must include 'total_ea', 'sku', 'sku_no_prefix').
    item_to_avail_qty : dict
        Mapping of str(Item No.) -> available quantity in warehouse.
    """
    if not item_to_avail_qty:
        return

    flagged = []
    for r in rows:
        total_ea = r.get("total_ea")
        if total_ea is None or total_ea == "" or total_ea <= 0:
            continue

        # Look up WH inventory by item number
        if r.get("is_bundle"):
            # For bundles: check each component
            for cd in r.get("component_details", []):
                item_no = cd.get("item_no")
                if item_no is None:
                    continue
                wh_inv = item_to_avail_qty.get(str(item_no).strip(), None)
                comp_ea = cd.get("comp_ea", 0)
                if wh_inv is not None and comp_ea > wh_inv:
                    flagged.append({
                        "sku": r["sku"],
                        "item_no": item_no,
                        "total_ea": comp_ea,
                        "wh_inv": int(wh_inv),
                        "deficit": int(comp_ea - wh_inv),
                        "is_component": True,
                    })
        else:
            item_no = extract_item_number(r.get("sku", ""))
            if item_no is None:
                continue
            wh_inv = item_to_avail_qty.get(str(item_no).strip(), None)
            if wh_inv is not None and total_ea > wh_inv:
                flagged.append({
                    "sku": r["sku"],
                    "item_no": item_no,
                    "total_ea": int(total_ea),
                    "wh_inv": int(wh_inv),
                    "deficit": int(total_ea - wh_inv),
                    "is_component": False,
                })

    if flagged:
        print(f"\n{'=' * 70}")
        print(f"  ⚠️  INVENTORY FLAG REPORT — Total EA > Current WH Inventory")
        print(f"  {len(flagged)} item(s) where shipment exceeds warehouse stock:")
        print(f"{'=' * 70}")
        print(f"  {'SKU':<30} {'Item#':<10} {'Total EA':>10} {'WH Inv':>10} {'Deficit':>10}")
        print(f"  {'-'*30} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")
        for f in sorted(flagged, key=lambda x: -x["deficit"]):
            comp_tag = " (component)" if f["is_component"] else ""
            print(f"  {f['sku']:<30} {str(f['item_no']):<10} "
                  f"{f['total_ea']:>10,} {f['wh_inv']:>10,} {f['deficit']:>10,}{comp_tag}")
        print(f"{'=' * 70}")
    else:
        print("\n  ✅ Inventory flag check: All items have sufficient WH inventory.")

    return flagged


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    """Entry point — orchestrates the full replenishment pipeline."""
    today = date.today().isoformat()

    print("=" * 58)
    print("  FBA Replenishment Automation")
    print(f"  Date : {today}")
    print(f"  Dir  : {os.path.abspath(CONFIG['input_dir'])}")
    print("=" * 58)

    # ---- 1. Load source files ----
    print("\nLocating source files ...")
    sources = load_sources()

    bts_df = sources["bts_df"]
    inv_col = sources["bts_inv_column"]
    item_df = sources["item_df"]
    data1_df = sources["data1_df"]
    data2_df = sources["data2_df"]
    instruction_df = sources["instruction_df"]
    bundles_df = sources["bundles_df"]

    # ---- 2. Filter rows with inventory to send ----
    removal_log: list[dict] = []
    dropped_inv = bts_df[bts_df[inv_col] <= 0]
    for _, row in dropped_inv.iterrows():
        sku_val = row.get("Merchant SKU")
        if pd.isna(sku_val) or str(sku_val).strip() in ("", "nan"):
            continue
        removal_log.append({"SKU": str(sku_val), "Reason": "Inv to send <= 0", "Value": row[inv_col]})
    bts_df = bts_df[bts_df[inv_col] > 0].copy()
    if CONFIG["sample_size"]:
        sample_dropped = bts_df.iloc[CONFIG["sample_size"]:]
        for _, row in sample_dropped.iterrows():
            sku_val = row.get("Merchant SKU")
            if pd.isna(sku_val) or str(sku_val).strip() in ("", "nan"):
                continue
            removal_log.append({"SKU": str(sku_val), "Reason": "Sample size limit", "Value": CONFIG["sample_size"]})
        bts_df = bts_df.head(CONFIG["sample_size"])

    print(f"\n  Rows with inventory > 0 : {len(bts_df)}")

    # ---- 3. Deduplicate BTS rows ----
    bts_dupes = bts_df.duplicated(subset=["Merchant SKU"], keep="first")
    if bts_dupes.any():
        for idx in bts_df.index[bts_dupes]:
            sku_val = bts_df.loc[idx, "Merchant SKU"]
            if pd.isna(sku_val) or str(sku_val).strip() in ("", "nan"):
                continue
            removal_log.append({"SKU": str(sku_val), "Reason": "Duplicate Merchant SKU", "Value": ""})
    bts_df = bts_df.drop_duplicates(subset=["Merchant SKU"], keep="first")

    # ---- 4. Split into bundles vs regular items ----
    bundle_mask = bts_df["Merchant SKU"].str.contains("+", na=False, regex=False)
    bundle_bts = bts_df[bundle_mask].copy()
    regular_bts = bts_df[~bundle_mask].copy()
    print(f"  Regular items           : {len(regular_bts)}")
    print(f"  Bundle items            : {len(bundle_bts)}")

    # ---- 5. Prepare Item List ----
    # Keep Item No. as-is (int or str) — variants like "942-A" must remain
    item_df = item_df.dropna(subset=["Item No."])
    item_df = item_df[item_df["Item No."].astype(str).str.strip() != ""]
    item_df = item_df.drop_duplicates(subset=["Item No."], keep="first")

    # ---- 6. Build UoM lookup from Data1 (col 0 = SKU, col 2 = UOM) ----
    sku_to_uom = dict(
        zip(
            data1_df.iloc[:, 0].astype(str),
            data1_df.iloc[:, 2].astype(str),
        )
    )

    rows: list[dict] = []
    missing_instructions: list[dict] = []
    missing_pack_data: list[str] = []

    # ---- 8. Process regular items first ----
    regular_bts["Item No."] = regular_bts["Merchant SKU"].apply(extract_item_number)
    dropped_item_no = regular_bts[regular_bts["Item No."].isna()]
    for _, row in dropped_item_no.iterrows():
        sku_val = row.get("Merchant SKU")
        if pd.isna(sku_val) or str(sku_val).strip() in ("", "nan"):
            continue
        removal_log.append({"SKU": str(sku_val), "Reason": "No Item No.", "Value": ""})
    regular_bts = regular_bts.dropna(subset=["Item No."])

    # Available Qty file has no Group Name — no discontinued filter
    item_df_active = item_df

    merged = regular_bts.merge(item_df_active, on="Item No.", how="left")

    rtype = CONFIG["replenish_type"]
    instr_col_name = "Instruction FBA" if rtype == "FBA" else "Instruction WFS"

    for _, row in merged.iterrows():
        sku = str(row["Merchant SKU"])
        sku_no_prefix = sku.replace("FBA_", "")
        uom_str = sku_to_uom.get(sku_no_prefix) or extract_suffix(sku) or "EACH"
        uom = parse_uom(uom_str)

        has_case_qty = pd.notna(row.get("Case Qty"))
        has_dims = pd.notna(row.get("Case Length"))
        if not has_case_qty:
            label = "(has dimensions but no Case Qty)" if has_dims else "(not in Item List)"
            missing_pack_data.append(f"{sku}  {label}")

        case_dims = {
            "length": row.get("Case Length", 0),
            "width": row.get("Case Width", 0),
            "height": row.get("Case Height", 0),
            "weight": row.get("Case Weight", 0),
        }

        box_dims = (
            {
                "length": row.get("Box Length", 0),
                "width": row.get("Box Width", 0),
                "height": row.get("Box Height", 0),
                "weight": row.get("Box Weight", 0),
            }
            if pd.notna(row.get("Box Qty"))
            else None
        )

        item_no = row["Item No."]
        # CASE_QTY_OVERRIDES uses int keys (8004, 8005, 8006); variants like "942-A" have no override
        try:
            item_no_int = int(float(str(item_no)))
        except (ValueError, TypeError):
            item_no_int = None
        case_qty_val = CASE_QTY_OVERRIDES.get(item_no_int, row.get("Case Qty", 1)) if item_no_int is not None else row.get("Case Qty", 1)
        box_qty_val = row.get("Box Qty")

        # Prefer Data2 values over Item List (same chain as WH formulas)
        d2_vals = lookup_pack_qty_from_data2(sku_no_prefix, data1_df, data2_df)
        if d2_vals["case_qty"] > 0:
            case_qty_val = CASE_QTY_OVERRIDES.get(item_no_int, d2_vals["case_qty"]) if item_no_int is not None else d2_vals["case_qty"]
        if d2_vals["box_qty"] > 0:
            box_qty_val = d2_vals["box_qty"]

        suffix_upper = extract_suffix(sku).upper()
        if suffix_upper == "CASE":
            if item_no_int is not None and item_no_int in CASE_QTY_OVERRIDES:
                # 8004/8005/8006: treat as singular (each), 5 per case, WH UoM = 1
                uom = 1
                result = select_pack_unit(
                    inv_to_send=row[inv_col],
                    case_qty=CASE_QTY_OVERRIDES[item_no_int],
                    box_qty=None,
                    uom=uom,
                    case_dims=case_dims,
                    box_dims=box_dims,
                )
                result["wh_uom"] = 1
                result["manifest_units_per_box"] = CASE_QTY_OVERRIDES[item_no_int]
                result["manifest_num_boxes"] = result["adj_sets"] // CASE_QTY_OVERRIDES[item_no_int]
            else:
                # Other CASE: no rounding, ship exactly inv_to_send cases; WH UoM = case_qty
                result = _build_case_only_result(
                    inv_to_send=row[inv_col],
                    case_qty=case_qty_val,
                    case_dims=case_dims,
                )
        elif suffix_upper == "BOX":
            # BOX: round to boxes_per_case, always ship cases; WH UoM = box_qty
            result = _build_box_suffix_result(
                inv_to_send=row[inv_col],
                case_qty=case_qty_val,
                box_qty=box_qty_val,
                case_dims=case_dims,
            )
        else:
            # EACH / SET{N} / PACK{N}: standard LCM rounding
            uom = parse_uom(sku_to_uom.get(sku_no_prefix) or extract_suffix(sku) or "EACH")
            result = select_pack_unit(
                inv_to_send=row[inv_col],
                case_qty=case_qty_val,
                box_qty=box_qty_val,
                uom=uom,
                case_dims=case_dims,
                box_dims=box_dims,
            )

        result["sku"] = sku
        result["sku_no_prefix"] = sku_no_prefix
        result["inv_to_send"] = int(_num(row[inv_col]))
        result["uom"] = result.get("wh_uom", result.get("uom", 1))
        result["is_bundle"] = False
        result["asin"] = (
            str(row["ASIN"]) if pd.notna(row.get("ASIN")) else ""
        )
        result["fnsku"] = (
            str(row["FNSKU"]) if pd.notna(row.get("FNSKU")) else ""
        )

        instr_text = get_instruction_from_sheet(sources["replenish_wb"], sku_no_prefix, rtype)
        needs_fallback = instr_text is None

        if needs_fallback:
            fallback = get_instruction_fallback(sku, instruction_df)
            if fallback:
                result["instruction_fallback"] = fallback
                missing_instructions.append({"sku": sku, **fallback})
            else:
                result["instruction_fallback"] = None
                print(f"  WARNING: no instruction match for {sku}")
        else:
            result["instruction_fallback"] = None
            result["instruction_text"] = instr_text

        rows.append(result)

    # ---- 9. Process bundle items (at end) ----
    for _, brow in bundle_bts.iterrows():
        sku = str(brow["Merchant SKU"])
        sku_no_prefix = sku.replace("FBA_", "")

        result = process_bundle(sku, brow[inv_col], item_df, bundles_df)
        result["sku"] = sku
        result["sku_no_prefix"] = sku_no_prefix
        result["inv_to_send"] = int(_num(brow[inv_col]))
        result["asin"] = (
            str(brow["ASIN"]) if pd.notna(brow.get("ASIN")) else ""
        )
        result["fnsku"] = (
            str(brow["FNSKU"]) if pd.notna(brow.get("FNSKU")) else ""
        )
        result["instruction_fallback"] = None
        result["uom"] = result.get("bundle_uom", 1)

        print(f"    BUNDLE {sku_no_prefix}: sellable_set={result['sellable_set']}, "
              f"inv={brow[inv_col]} -> adj={result['adj_sets']}, "
              f"packs={result['num_packs']}")
        rows.append(result)

    if not rows:
        print("\nNo items to process — exiting.")
        return

    if removal_log:
        print(f"\n  Removed items: {len(removal_log)} (see 'Removed Items' sheet in WH output)")

    # ---- 7b. Round Actual Qty to unit multiple (Fix 4: 17090-A-EACH etc.) ----
    unit_rounding_flags = []
    for r in rows:
        if r.get("is_bundle"):
            continue
        actual = r.get("adj_sets", 0)
        pack_type = r.get("pack_type", "")
        pack_qty = r.get("pack_qty", 0)
        uom = r.get("uom", 1) or 1
        unit_qty = math.floor(pack_qty / uom) if uom and pack_qty else 0
        if unit_qty <= 0:
            unit_qty = 1
        rounded, flag = round_to_unit_multiple(actual, unit_qty, r.get("sku", "?"))
        if rounded != actual:
            r["adj_sets"] = int(rounded)
            r["total_ea"] = int(rounded * uom)
            upb = r.get("manifest_units_per_box") or r.get("sets_per_pack")
            if upb and upb > 0:
                r["manifest_num_boxes"] = math.ceil(rounded / upb)
            r["num_packs"] = r.get("manifest_num_boxes", r.get("num_packs", 1))
        if flag:
            unit_rounding_flags.append(flag)
    if unit_rounding_flags:
        print("\n  Actual Qty rounding (below minimum pack):")
        for f in unit_rounding_flags:
            print(f"    {f}")

    # ---- 7c. Sort: high inventory first, bundles last ----
    avail_col = None
    for c in ["Avail. Qty", "Available Qty", "Avail Qty"]:
        if c in item_df.columns:
            avail_col = c
            break
    item_to_avail_qty = {}
    if avail_col:
        for _, r in item_df.iterrows():
            ino = r.get("Item No.")
            if pd.notna(ino):
                qty = _num(r.get(avail_col), 0)
                key = str(ino).strip()
                if key not in item_to_avail_qty or qty > item_to_avail_qty[key]:
                    item_to_avail_qty[key] = qty

    def _avail_for_row(r):
        if r.get("is_bundle"):
            comps = r.get("component_details", [])
            if not comps:
                return 0
            qtys = [item_to_avail_qty.get(str(c.get("item_no", "")).strip(), 0) for c in comps]
            return min(qtys) if qtys else 0
        item_no = extract_item_number(r.get("sku", ""))
        return item_to_avail_qty.get(str(item_no).strip(), 0) if item_no is not None else 0

    rows.sort(key=lambda r: (r.get("is_bundle", False), _avail_for_row(r) < 100))

    # ---- 8. Write output files ----
    manifest_path = unique_path(os.path.join(
        CONFIG["output_dir"],
        CONFIG["manifest_output"].format(date=today),
    ))
    replenish_path = unique_path(os.path.join(
        CONFIG["output_dir"],
        CONFIG["replenish_output"].format(date=today),
    ))

    # ---- 8a. Update Instruction sheet with fallback values ----
    update_instruction_sheet(
        sources["replenish_wb"], rows, instruction_df, CONFIG["replenish_type"]
    )

    # ---- 8b. Auto-populate Data1 / Data2 with missing entries ----
    update_data1_sheet(sources["replenish_wb"], rows)
    update_data2_sheet(sources["replenish_wb"], rows, item_df)

    print("\nWriting output files ...")
    validate_manifest_num_boxes(rows)
    # Write WH first so recalc can populate cached formula values for FBA
    write_replenishment(
        sources["replenish_wb"],
        rows,
        replenish_path,
        removal_log=removal_log,
        data1_df=sources["data1_df"],
        data2_df=sources["data2_df"],
        item_to_avail_qty=item_to_avail_qty,
    )

    wh_data = None
    if CONFIG.get("run_recalc", True):
        wh_data = _recalc_and_read_wh(replenish_path, len(rows))
        if wh_data is None:
            print("  NOTE: recalc skipped or failed — FBA using Python-computed values")
        else:
            print("  Recalculated WH formulas — FBA using cached WH values")

    write_manifest(
        sources["manifest_wb"],
        rows,
        manifest_path,
        wh_data=wh_data,
        data1_df=sources["data1_df"],
        data2_df=sources["data2_df"],
    )

    # ---- 9. Summary ----
    print_summary(rows, missing_instructions, missing_pack_data)

    # ---- 10. Inventory flag report (Total EA vs Current WH Inventory) ----
    print_inventory_flag_report(rows, item_to_avail_qty)

    print(f"  Manifest  -> {os.path.abspath(manifest_path)}")
    print(f"  Replenish -> {os.path.abspath(replenish_path)}")
    print("\nDone.")


if __name__ == "__main__":
    main()
