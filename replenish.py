"""FBA Replenishment Automation Script.

Reads BTS Calcs and Item List workbooks, computes case-vs-box pack-unit
quantities, and produces two output files:

  1. Amazon FBA Manifest  (Send-to-Amazon workflow upload)
  2. Warehouse Replenishment Sheet  (packing sheet for the warehouse)

Dependencies: pandas, openpyxl
Usage:        Place source files in the working directory and run:
              python replenish.py
"""

import math
import os
import re
from datetime import date

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# ---------------------------------------------------------------------------
# Configuration — edit these if your folder layout differs
# ---------------------------------------------------------------------------

CONFIG = {
    "input_dir": ".",
    "output_dir": ".",
    "manifest_output": "FBA_Manifest_{date}.xlsx",
    "replenish_output": "WH_Replenishment_{date}.xlsx",
    "sample_size": None,  # None = process all rows where inv > 0
    "replenish_type": "FBA",  # "FBA" or "WFS" — controls which instruction column to populate
}

# Items whose Case Qty must be forced to a specific value,
# regardless of what the Item List spreadsheet contains.
CASE_QTY_OVERRIDES = {
    8004: 5,
    8005: 5,
    8006: 5,
}

# Alternating row colours for readability
FILL_LIGHT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _num(value, default=0.0):
    """Safely coerce *value* to float, returning *default* for NaN / None / non-numeric."""
    if value is None:
        return default
    try:
        f = float(value)
        return default if math.isnan(f) else f
    except (TypeError, ValueError):
        return default

# ---------------------------------------------------------------------------
# Output path helper
# ---------------------------------------------------------------------------


def unique_path(path):
    """Return *path* if it doesn't exist; otherwise append _v2, _v3, … until free."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    version = 2
    while True:
        candidate = f"{base}_v{version}{ext}"
        if not os.path.exists(candidate):
            return candidate
        version += 1


# ---------------------------------------------------------------------------
# File scanning
# ---------------------------------------------------------------------------


def find_file(keywords, directory=None):
    """Locate a single .xlsx/.xls file whose name contains *all* given keywords.

    Parameters
    ----------
    keywords : str or list[str]
        Substrings that must ALL appear in the filename (case-insensitive).
        Underscores in a keyword also match the equivalent space.
    directory : str, optional
        Folder to scan.  Defaults to ``CONFIG["input_dir"]``.

    Returns
    -------
    str
        Full path to the matched file.

    Raises
    ------
    FileNotFoundError
        When no file matches the keywords.
    """
    if isinstance(keywords, str):
        keywords = [keywords]
    directory = directory or CONFIG["input_dir"]

    def _matches(filename):
        fn = filename.lower()
        for kw in keywords:
            lo = kw.lower()
            alt = kw.replace("_", " ").lower()
            if lo not in fn and alt not in fn:
                return False
        return True

    candidates = [
        f
        for f in os.listdir(directory)
        if f.lower().endswith((".xlsx", ".xls"))
        and not f.startswith("~$")
        and _matches(f)
    ]

    if not candidates:
        raise FileNotFoundError(
            f"No Excel file matching {keywords} found in "
            f"{os.path.abspath(directory)}"
        )

    if len(candidates) == 1:
        return os.path.join(directory, candidates[0])

    print(f"\nMultiple files match {keywords}:")
    for i, name in enumerate(candidates, 1):
        print(f"  {i}. {name}")
    while True:
        try:
            choice = int(input("Enter number: ")) - 1
            if 0 <= choice < len(candidates):
                return os.path.join(directory, candidates[choice])
        except ValueError:
            pass
        print("Invalid choice — try again.")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_sources():
    """Read every source workbook and return DataFrames + openpyxl workbooks.

    Returns a dict with keys:
        bts_df, item_df, data1_df, data2_df, instruction_df, bundles_df,
        replenish_wb, manifest_wb
    """
    bts_path = find_file("BTS_Calcs")
    item_path = find_file("Item_List")
    replenish_tpl_path = find_file(["Replenishment", "FBA"])
    manifest_tpl_path = find_file("ManifestFileUpload")

    print(f"  BTS Calcs     : {os.path.basename(bts_path)}")
    print(f"  Item List     : {os.path.basename(item_path)}")
    print(f"  Replenish Tpl : {os.path.basename(replenish_tpl_path)}")
    print(f"  Manifest Tpl  : {os.path.basename(manifest_tpl_path)}")

    bts_df = pd.read_excel(bts_path, sheet_name="Working Sheet")
    item_df = pd.read_excel(item_path, sheet_name="Sheet1")

    replenish_wb = load_workbook(replenish_tpl_path)
    data1_df = pd.read_excel(replenish_tpl_path, sheet_name="Data1")
    data2_df = pd.read_excel(replenish_tpl_path, sheet_name="Data2")
    instruction_df = pd.read_excel(replenish_tpl_path, sheet_name="Instruction")
    bundles_df = pd.read_excel(replenish_tpl_path, sheet_name="Bundles")

    manifest_wb = load_workbook(manifest_tpl_path)

    return {
        "bts_df": bts_df,
        "item_df": item_df,
        "data1_df": data1_df,
        "data2_df": data2_df,
        "instruction_df": instruction_df,
        "bundles_df": bundles_df,
        "replenish_wb": replenish_wb,
        "manifest_wb": manifest_wb,
    }


# ---------------------------------------------------------------------------
# Item-number extraction
# ---------------------------------------------------------------------------


def extract_item_number(sku):
    """Return the leading integer from an SKU after stripping any ``FBA_`` prefix.

    Examples::

        FBA_8006-CASE      -> 8006
        FBA_2227-EACH-UPC  -> 2227
        510-EACH-B         -> 510
    """
    if pd.isna(sku) or sku is None:
        return None
    clean = str(sku).replace("FBA_", "")
    m = re.match(r"(\d+)", clean)
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# UoM parsing
# ---------------------------------------------------------------------------


def parse_uom(uom_str):
    """Convert a UoM string to the number of pieces per sellable unit.

    ``SET6`` -> 6,  ``EACH`` -> 1,  ``PACK2`` -> 2.
    Returns 1 when the string contains no digits.
    """
    if not uom_str or pd.isna(uom_str):
        return 1
    digits = re.search(r"\d+", str(uom_str))
    return int(digits.group()) if digits else 1


# ---------------------------------------------------------------------------
# Bundle detection & parsing
# ---------------------------------------------------------------------------


def is_bundle(sku):
    """True if *sku* contains ``+``, indicating a multi-item bundle."""
    return "+" in str(sku)


def round_to_unit_multiple(actual, unit_qty, sku):
    """Round Actual Qty to nearest multiple of unit_qty (CASE or BOX); flag if gap > 20%.

    unit_qty = CASE qty for Pack Unit=CASE, BOX qty for Pack Unit=BOX.
    Ensures we never ship partial packs (e.g. 12 when min BOX is 24).
    """
    if not unit_qty or unit_qty == 0:
        return actual, None
    remainder = actual % unit_qty
    if remainder == 0:
        return actual, None
    lower = actual - remainder
    upper = lower + unit_qty
    rounded = lower if (actual - lower) <= (upper - actual) else upper
    gap_pct = abs(rounded - actual) / actual if actual > 0 else 0
    flag = None
    if gap_pct > 0.20:
        flag = (
            f"⚠️ SKU {sku}: Actual qty {actual} rounded to {rounded} "
            f"(gap: {gap_pct:.0%}) — please recheck recommended qty"
        )
    return rounded, flag


def round_to_case_multiple(recommended, units_per_box, sku):
    """Round recommended qty to nearest multiple of units_per_box; flag if gap > 20%."""
    if units_per_box is None or units_per_box == 0:
        return recommended, None
    lower = math.floor(recommended / units_per_box) * units_per_box
    upper = math.ceil(recommended / units_per_box) * units_per_box
    diff_lower = abs(recommended - lower)
    diff_upper = abs(recommended - upper)
    closest = lower if diff_lower <= diff_upper else upper
    gap_pct = abs(closest - recommended) / recommended if recommended > 0 else 0
    flag = None
    if gap_pct > 0.20:
        flag = (
            f"⚠️ SKU {sku}: Recommended qty {recommended} rounded to {closest} "
            f"(gap: {gap_pct:.0%}) — please recheck"
        )
    return closest, flag


def resolve_bundle_case_qty(sku, uom, item_list_df, pack_unit):
    """Resolve Units per box for bundle SKUs from ItemList.

    Bundle SKUs have no CASE/BOX in For WH — look up each component's
    Case Qty or Box Qty in ItemList.

    For equal-ratio bundles (e.g. 534-EA+535-EA): floor(min(qty_list) / uom).
    For weighted-ratio bundles (e.g. 803-SET4+804-SET6): extract ratio from
    -SETn suffix, scale box counts by ratio, total_units = sum(box_count_i * box_qty_i),
    units_per_box = floor(total_units / uom).

    Parameters
    ----------
    sku : str
        Bundle SKU (e.g. 534-EA+535-EA or 803-SET4+804-SET6).
    uom : int
        Pieces per sellable set.
    item_list_df : pandas.DataFrame
        ItemList with columns "Item No.", "Case Qty", "Box Qty".
    pack_unit : str
        "CASE" or "BOX" — which quantity column to use.

    Returns
    -------
    int or None
        Units per box, or None if not a bundle or any component not in ItemList.
    """
    if "+" not in str(sku):
        return None
    parts = str(sku).replace("FBA_", "").split("+")
    qty_col = "Case Qty" if pack_unit == "CASE" else "Box Qty"
    items = []  # list of (item_no, qty, ratio)
    for part in parts:
        part = part.strip()
        m = re.match(r"^(\d+)", part)
        item_no = int(m.group(1)) if m else None
        if item_no is None:
            continue
        ratio_m = re.search(r"-SET(\d+)", part, re.IGNORECASE)
        ratio = int(ratio_m.group(1)) if ratio_m else 1
        row = item_list_df[item_list_df["Item No."] == item_no]
        if row.empty:
            print(f"  ⚠️ Bundle SKU {sku}: component item {item_no} not found in ItemList — "
                  "Units per box could not be calculated. Please add item to ItemList.")
            return None
        qty = _num(row.iloc[0].get(qty_col), 0)
        if qty <= 0:
            continue
        items.append((item_no, qty, ratio))
    if not items:
        return None
    ratios = [r for (_, _, r) in items]
    if any(r != 1 for r in ratios):
        g = math.gcd(ratios[0], ratios[1]) if len(ratios) >= 2 else ratios[0]
        for r in ratios[2:]:
            g = math.gcd(g, r)
        box_counts = [r // g for r in ratios]
        total_units = sum(bc * qty for (_, qty, _), bc in zip(items, box_counts))
    else:
        total_units = min(qty for (_, qty, _) in items)
    if uom <= 0:
        return int(total_units)
    return math.floor(total_units / uom)


def parse_bundle_components(sku):
    """Parse a bundle SKU into its component parts.

    Parameters
    ----------
    sku : str
        Full SKU such as ``FBA_534-EA+535-EA`` or ``534-EA+535-EA``.

    Returns
    -------
    list[dict]
        One dict per component with keys ``raw``, ``item_no``, ``ea_per_set``.

    Example::

        parse_bundle_components("FBA_534-EA+535-EA")
        # [{"raw": "534-EA", "item_no": 534, "ea_per_set": 1},
        #  {"raw": "535-EA", "item_no": 535, "ea_per_set": 1}]
    """
    clean = str(sku).replace("FBA_", "")
    parts = clean.split("+")
    components = []
    for part in parts:
        item_no = extract_item_number(part)
        uom_suffix = extract_suffix("FBA_" + part)
        ea = parse_uom(uom_suffix)
        components.append({"raw": part, "item_no": item_no, "ea_per_set": ea})
    return components


# ---------------------------------------------------------------------------
# Pack-unit selection
# ---------------------------------------------------------------------------


def _build_case_only_result(inv_to_send, case_qty, case_dims):
    """Build result for CASE suffix (non-8004/5/6): no rounding, ship exactly inv_to_send cases.

    Each case is 1 sellable unit. WH UoM = case_qty. Manifest: Units per box = 1, Num boxes = adj_qty.
    """
    adj_sets = max(1, int(_num(inv_to_send)))
    num_packs = adj_sets
    cq = int(_num(case_qty, 1))
    total_ea = adj_sets * cq
    dims = case_dims or {}
    return {
        "pack_type": "CASE",
        "pack_qty": cq,
        "sets_per_pack": 1,
        "adj_sets": adj_sets,
        "total_ea": total_ea,
        "num_packs": num_packs,
        "length": math.floor(_num(dims.get("length"))),
        "width": math.floor(_num(dims.get("width"))),
        "height": math.floor(_num(dims.get("height"))),
        "weight": round(_num(dims.get("weight")), 2),
        "wh_uom": cq,
        "manifest_units_per_box": 1,
        "manifest_num_boxes": adj_sets,
        "remainder_flag": None,
    }


def _build_box_suffix_result(inv_to_send, case_qty, box_qty, case_dims):
    """Build result for BOX suffix: round to boxes_per_case, always ship full cases.

    Sellable unit = 1 box. WH UoM = box_qty. Manifest: Units per box = case_qty/box_qty,
    Num boxes = physical cases shipped.
    """
    inv_to_send = _num(inv_to_send)
    cq = int(_num(case_qty, 1))
    bq = int(_num(box_qty, 1))
    if bq <= 0:
        bq = 1
    boxes_per_case = cq // bq
    if boxes_per_case <= 0:
        boxes_per_case = 1

    # Round inv_to_send to nearest multiple of boxes_per_case
    lo = (math.floor(inv_to_send / boxes_per_case)) * boxes_per_case
    hi = lo + boxes_per_case
    if abs(hi - inv_to_send) <= abs(inv_to_send - lo):
        adj_sets = int(hi)
    else:
        adj_sets = int(lo)
    adj_sets = max(adj_sets, boxes_per_case)  # minimum 1 full case

    num_packs = adj_sets // boxes_per_case  # physical cases
    total_ea = adj_sets * bq  # total items (EA)
    dims = case_dims or {}
    return {
        "pack_type": "CASE",
        "pack_qty": cq,
        "sets_per_pack": boxes_per_case,
        "adj_sets": adj_sets,
        "total_ea": total_ea,
        "num_packs": num_packs,
        "length": math.floor(_num(dims.get("length"))),
        "width": math.floor(_num(dims.get("width"))),
        "height": math.floor(_num(dims.get("height"))),
        "weight": round(_num(dims.get("weight")), 2),
        "wh_uom": bq,
        "manifest_units_per_box": boxes_per_case,
        "manifest_num_boxes": num_packs,
        "remainder_flag": None,
    }


def select_pack_unit(inv_to_send, case_qty, box_qty, uom, case_dims, box_dims):
    """Decide CASE vs BOX and compute adjusted quantities and dimensions.

    Parameters
    ----------
    inv_to_send : float
        Number of **sellable sets** to ship (from BTS Calcs).
        NOT individual EA — multiply by *uom* to get EA.
    case_qty : float
        Individual pieces (EA) per case, from Item List.
    box_qty : float or None
        Individual pieces (EA) per box, from Item List.
    uom : int
        Pieces per sellable set (e.g. SET24 → 24, EACH → 1).
    case_dims, box_dims : dict or None
        Length / width / height / weight for the pack unit.

    Decision rules (all comparisons in EA)
    ---------------------------------------
    total_ea_raw = inv_to_send * uom
    threshold    = 0.8 * case_qty

    * total_ea_raw >= threshold             -> CASE
    * total_ea_raw <  threshold AND box_qty -> BOX  (two-pass: if BOX-rounded
      total_ea >= threshold, switch to CASE instead)
    * total_ea_raw <  threshold AND no box  -> CASE  (minimum 1 full case)

    If pack_qty % uom != 0 (partial-pack): skip LCM rounding, keep
    adj_sets = inv_to_send, and use math.ceil for num_packs.  The
    warehouse pulls enough packs and sets aside leftover pieces.
    """
    inv_to_send = _num(inv_to_send)          # sellable sets
    case_qty = _num(case_qty, default=1.0)   # EA per case
    if case_qty == 0:
        case_qty = 1.0
    uom = max(1, int(_num(uom, default=1)))  # EA per set
    has_box = _num(box_qty) > 0

    total_ea_raw = inv_to_send * uom         # convert sets → EA
    threshold = 0.8 * case_qty

    if total_ea_raw >= threshold or not has_box:
        # --- CASE path ---
        cq = int(case_qty)
        sets_per_pack = math.floor(cq / uom) if uom > 0 else cq
        if sets_per_pack <= 0:
            sets_per_pack = 1
        remainder_flag = f"⚠️ Remainder: {cq % uom} unit(s) per case will not form a complete set" if (cq % uom) > 0 else None

        if cq % uom != 0:
            # Partial-pack: pack_qty and UoM are misaligned — keep original
            # quantity and pull enough cases, setting aside leftovers.
            num_packs = max(1, math.ceil(inv_to_send / sets_per_pack))
            adj_sets = sets_per_pack * num_packs
            total_ea = cq * num_packs
        else:
            # Round inv_to_send to nearest multiple of sets_per_pack
            case_lo = max(1, math.floor(inv_to_send / sets_per_pack))
            case_hi = math.ceil(inv_to_send / sets_per_pack)
            if case_hi < 1:
                case_hi = 1
            lo_sets = case_lo * sets_per_pack
            hi_sets = case_hi * sets_per_pack
            if abs(hi_sets - inv_to_send) <= abs(inv_to_send - lo_sets):
                num_packs = case_hi
            else:
                num_packs = case_lo
            adj_sets = sets_per_pack * num_packs
            total_ea = cq * num_packs

        dims = case_dims
        pack_type = "CASE"
        pack_qty = case_qty
    else:
        # --- BOX path ---
        bq = int(float(box_qty))
        sets_per_pack = math.floor(bq / uom) if uom > 0 else bq
        if sets_per_pack <= 0:
            sets_per_pack = 1
        remainder_flag = f"⚠️ Remainder: {bq % uom} unit(s) per box will not form a complete set" if (bq % uom) > 0 else None

        if bq % uom != 0:
            # Partial-pack for BOX
            num_packs = max(1, math.ceil(inv_to_send / sets_per_pack))
            adj_sets = sets_per_pack * num_packs
            total_ea = bq * num_packs
        else:
            case_lo = max(1, math.floor(inv_to_send / sets_per_pack))
            case_hi = math.ceil(inv_to_send / sets_per_pack)
            if case_hi < 1:
                case_hi = 1
            lo_sets = case_lo * sets_per_pack
            hi_sets = case_hi * sets_per_pack
            if abs(hi_sets - inv_to_send) <= abs(inv_to_send - lo_sets):
                num_packs = case_hi
            else:
                num_packs = case_lo
            adj_sets = sets_per_pack * num_packs
            total_ea = bq * num_packs

        # Two-pass: if the BOX-rounded quantity now meets the CASE threshold,
        # switch to CASE — shipping one case is more efficient than many boxes.
        if total_ea >= threshold:
            cq = int(case_qty)
            sets_per_pack = math.floor(cq / uom) if uom > 0 else cq
            if sets_per_pack <= 0:
                sets_per_pack = 1
            remainder_flag = f"⚠️ Remainder: {cq % uom} unit(s) per case will not form a complete set" if (cq % uom) > 0 else None
            if cq % uom != 0:
                num_packs = max(1, math.ceil(inv_to_send / sets_per_pack))
                adj_sets = sets_per_pack * num_packs
                total_ea = cq * num_packs
            else:
                case_lo = max(1, math.floor(inv_to_send / sets_per_pack))
                case_hi = math.ceil(inv_to_send / sets_per_pack)
                if case_hi < 1:
                    case_hi = 1
                lo_sets = case_lo * sets_per_pack
                hi_sets = case_hi * sets_per_pack
                if abs(hi_sets - inv_to_send) <= abs(inv_to_send - lo_sets):
                    num_packs = case_hi
                else:
                    num_packs = case_lo
                adj_sets = sets_per_pack * num_packs
                total_ea = cq * num_packs
            dims = case_dims
            pack_type = "CASE"
            pack_qty = case_qty
        else:
            num_packs = total_ea // bq if bq > 0 else 1
            dims = box_dims if box_dims else case_dims
            pack_type = "BOX"
            pack_qty = float(box_qty)
    # For EACH/SET/PACK: WH UoM = uom, manifest units/box = sets per pack, num boxes = num_packs
    return {
        "pack_type": pack_type,
        "pack_qty": int(pack_qty),
        "sets_per_pack": sets_per_pack,
        "adj_sets": int(adj_sets),
        "total_ea": total_ea,
        "num_packs": num_packs,
        "length": math.floor(_num(dims.get("length"))),
        "width": math.floor(_num(dims.get("width"))),
        "height": math.floor(_num(dims.get("height"))),
        "weight": round(_num(dims.get("weight")), 2),
        "wh_uom": uom,
        "manifest_units_per_box": sets_per_pack,
        "manifest_num_boxes": num_packs,
        "remainder_flag": remainder_flag,
    }


# ---------------------------------------------------------------------------
# Bundle: Sellable-Set calculation + processing
# ---------------------------------------------------------------------------


def calculate_sellable_set(components, item_df, inv_to_send):
    """Compute the minimum bundle-set step that yields full packs for every component.

    For each component the pack unit (CASE vs BOX) is chosen using the same
    0.8-threshold rule as regular items.  The per-component minimum step is
    ``pack_qty / gcd(pack_qty, ea_per_set)``.  The overall Sellable Set is
    the LCM of all per-component steps.

    Returns
    -------
    sellable_set : int
    comp_info : list[dict]
        Per-component details: item_no, ea_per_set, pack_qty, pack_type, dims, raw.
    """
    comp_info = []
    min_sets_values = []

    for comp in components:
        item_no = comp["item_no"]
        ea_per_set = comp["ea_per_set"]
        item_row = item_df[item_df["Item No."] == item_no]

        case_qty = 1
        box_qty = 0
        case_dims = {"length": 0, "width": 0, "height": 0, "weight": 0}
        box_dims = None

        if not item_row.empty:
            r = item_row.iloc[0]
            case_qty = int(_num(
                CASE_QTY_OVERRIDES.get(item_no, r.get("Case Qty", 1)), 1
            ))
            if case_qty == 0:
                case_qty = 1
            box_qty = _num(r.get("Box Qty", 0))
            case_dims = {
                "length": r.get("Case Length", 0),
                "width": r.get("Case Width", 0),
                "height": r.get("Case Height", 0),
                "weight": r.get("Case Weight", 0),
            }
            if box_qty > 0:
                box_dims = {
                    "length": r.get("Box Length", 0),
                    "width": r.get("Box Width", 0),
                    "height": r.get("Box Height", 0),
                    "weight": r.get("Box Weight", 0),
                }

        comp_ea_raw = inv_to_send * ea_per_set
        threshold = 0.8 * case_qty
        if comp_ea_raw >= threshold or box_qty <= 0:
            pack_qty = case_qty
            pack_type = "CASE"
            dims = case_dims
        else:
            pack_qty = int(box_qty)
            pack_type = "BOX"
            dims = box_dims or case_dims

        min_sets_i = pack_qty // math.gcd(pack_qty, ea_per_set)
        min_sets_values.append(min_sets_i)

        comp_info.append({
            "item_no": item_no,
            "ea_per_set": ea_per_set,
            "pack_qty": pack_qty,
            "pack_type": pack_type,
            "dims": dims,
            "raw": comp["raw"],
        })

    sellable_set = min_sets_values[0]
    for ms in min_sets_values[1:]:
        sellable_set = (sellable_set * ms) // math.gcd(sellable_set, ms)

    return sellable_set, comp_info


def process_bundle(sku, inv_to_send, item_df, bundles_df):
    """Process a bundle SKU and return a result dict matching ``select_pack_unit`` output.

    Steps
    -----
    1. Parse components from *sku*.
    2. Calculate Sellable Set from component pack quantities.
    3. Round *inv_to_send* (bundle sets) to nearest multiple of Sellable Set
       (closest; ties round up).
    4. Compute per-component packs and EA.
    5. Aggregate into a single-row result.
    """
    components = parse_bundle_components(sku)
    inv_to_send = _num(inv_to_send)
    sellable_set, comp_info = calculate_sellable_set(
        components, item_df, inv_to_send
    )
    bundle_uom = sum(ci["ea_per_set"] for ci in comp_info)
    pack_types = {ci["pack_type"] for ci in comp_info}
    overall_pack_type = "CASE" if "CASE" in pack_types else "BOX"
    units_per_box = resolve_bundle_case_qty(sku, bundle_uom, item_df, overall_pack_type)

    if units_per_box and units_per_box > 0:
        adj_sets, rounding_flag = round_to_case_multiple(inv_to_send, units_per_box, sku)
        adj_sets = max(1, int(adj_sets))
    else:
        lo = max(sellable_set, (math.floor(inv_to_send / sellable_set)) * sellable_set)
        hi = lo + sellable_set
        if abs(hi - inv_to_send) <= abs(inv_to_send - lo):
            adj_sets = int(hi)
        else:
            adj_sets = int(lo)
        rounding_flag = None

    total_packs = 0
    total_ea_all = 0
    component_details = []
    for ci in comp_info:
        comp_ea = adj_sets * ci["ea_per_set"]
        comp_packs = comp_ea // ci["pack_qty"]
        total_packs += comp_packs
        total_ea_all += comp_ea
        component_details.append({**ci, "comp_ea": comp_ea, "comp_packs": comp_packs})

    best = max(component_details, key=lambda c: (
        _num(c["dims"].get("length"))
        * _num(c["dims"].get("width"))
        * _num(c["dims"].get("height"))
    ))

    sets_per_pack = adj_sets // total_packs if total_packs > 0 else adj_sets
    if units_per_box is None or units_per_box <= 0:
        units_per_box = sets_per_pack
    manifest_units_per_box = units_per_box if units_per_box is not None else sets_per_pack
    manifest_num_boxes = (
        math.ceil(adj_sets / manifest_units_per_box)
        if manifest_units_per_box and manifest_units_per_box > 0
        else total_packs
    )
    return {
        "pack_type": overall_pack_type,
        "pack_qty": best["pack_qty"],
        "sets_per_pack": sets_per_pack,
        "adj_sets": adj_sets,
        "total_ea": total_ea_all,
        "num_packs": total_packs,
        "manifest_units_per_box": manifest_units_per_box,
        "manifest_num_boxes": manifest_num_boxes,
        "length": math.floor(_num(best["dims"].get("length"))),
        "width": math.floor(_num(best["dims"].get("width"))),
        "height": math.floor(_num(best["dims"].get("height"))),
        "weight": round(_num(best["dims"].get("weight")), 2),
        "is_bundle": True,
        "bundle_uom": bundle_uom,
        "sellable_set": sellable_set,
        "component_details": component_details,
        "rounding_flag": rounding_flag,
    }


# ---------------------------------------------------------------------------
# Instruction fallback
# ---------------------------------------------------------------------------


def extract_suffix(sku):
    """Extract the UoM-like suffix from an FBA SKU.

    Scans hyphen-delimited tokens for a recognised UoM pattern so that
    variant indicators (e.g. the ``A`` in ``17090-A-EACH``) are skipped.

    ``FBA_2227-EACH-UPC`` -> ``EACH``,  ``FBA_582-SET6`` -> ``SET6``,
    ``FBA_17090-A-EACH`` -> ``EACH``,  ``FBA_5077-A-CASE`` -> ``CASE``
    """
    clean = str(sku).replace("FBA_", "")
    stripped = re.sub(r"^\d+", "", clean).lstrip("-")
    if not stripped:
        return ""
    tokens = stripped.split("-")
    _UOM_RE = re.compile(r"^(EACH|EA|SET\d*|PACK\d*|CASE|BOX)$", re.IGNORECASE)
    for token in tokens:
        if _UOM_RE.match(token):
            return token
    return tokens[0]


def get_instruction_fallback(sku, instruction_df):
    """Match a packing instruction by UoM suffix when the SKU is missing.

    Looks for rows in the Instruction sheet whose ``FBA SKU`` column ends
    with the same suffix as *sku*.

    Returns
    -------
    dict or None
        ``{instruction_fba, instruction_wfs, suffix_used}`` on match,
        ``None`` otherwise.
    """
    suffix = extract_suffix(sku)
    if not suffix:
        return None

    fba_col = instruction_df["FBA SKU"].astype(str).str.upper()
    hits = instruction_df.loc[fba_col.str.endswith(suffix.upper())]
    if hits.empty:
        return None

    row = hits.iloc[0]
    return {
        "instruction_fba": row.get("Instruction FBA", ""),
        "instruction_wfs": row.get("Instruction WFS", ""),
        "suffix_used": suffix,
    }


# ---------------------------------------------------------------------------
# Instruction sheet updater
# ---------------------------------------------------------------------------

# Instruction sheet column layout:
#   A=SKU  B=ASIN  C=FNSKU  D=Item  E=FBA SKU
#   F=Instruction FBA  G=AWD  H=Instruction WFS  I=GTIN  J=Supplies
INSTR_COL = {"FBA": 6, "WFS": 8}    # which column to write for each type


def update_instruction_sheet(replenish_wb, rows, instruction_df, replenish_type):
    """Add or update rows in the Instruction sheet for missing/blank instructions.

    * SKU not in sheet → append a new row with the fallback instruction.
    * SKU in sheet but instruction column for *replenish_type* is blank/0 →
      fill in the fallback instruction value.
    """
    ws = replenish_wb["Instruction"]
    col_idx = INSTR_COL.get(replenish_type, 6)

    # Build SKU → Excel row map from the existing Instruction sheet
    sku_row_map = {}
    for excel_row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=excel_row, column=1).value
        if cell_val:
            sku_row_map[str(cell_val).strip()] = excel_row

    next_row = ws.max_row + 1
    added = 0
    updated = 0

    for r in rows:
        fb = r.get("instruction_fallback")
        if not fb:
            continue
        instr_text = fb.get("instruction_fba") if replenish_type == "FBA" else fb.get("instruction_wfs")
        if not instr_text or (isinstance(instr_text, float) and math.isnan(instr_text)):
            continue

        sku_no_prefix = r["sku_no_prefix"]

        if sku_no_prefix in sku_row_map:
            # SKU exists — check if instruction cell is blank/0
            existing_row = sku_row_map[sku_no_prefix]
            existing_val = ws.cell(row=existing_row, column=col_idx).value
            if not existing_val or existing_val == 0:
                ws.cell(row=existing_row, column=col_idx, value=instr_text)
                updated += 1
        else:
            # SKU missing — add a new row
            ws.cell(row=next_row, column=1, value=sku_no_prefix)          # A  SKU
            ws.cell(row=next_row, column=2, value=r.get("asin", ""))      # B  ASIN
            ws.cell(row=next_row, column=3, value=r.get("fnsku", ""))     # C  FNSKU
            item_no = extract_item_number(r["sku"])
            ws.cell(row=next_row, column=4, value=item_no)                # D  Item
            ws.cell(row=next_row, column=5, value=r["sku"])               # E  FBA SKU
            ws.cell(row=next_row, column=col_idx, value=instr_text)       # F or H
            sku_row_map[sku_no_prefix] = next_row
            next_row += 1
            added += 1

    if added or updated:
        print(f"  Instruction sheet: {added} new row(s), {updated} updated row(s)")


# ---------------------------------------------------------------------------
# Data1 / Data2 sheet auto-population
# ---------------------------------------------------------------------------


def _detect_col_type(ws, column, start_row=2, sample=20):
    """Return ``'str'`` if the first non-empty values in *column* are strings, else ``'num'``."""
    for excel_row in range(start_row, min(start_row + sample, ws.max_row + 1)):
        val = ws.cell(row=excel_row, column=column).value
        if val is not None:
            return "str" if isinstance(val, str) else "num"
    return "num"


def update_data1_sheet(replenish_wb, rows):
    """Append rows to the Data1 sheet for any SKUs not already present.

    Data1 layout: A=SKU, B=ITEM (item number), C=UOM (numeric EA per set).
    UoM comes from the result dict (already resolved for CASE/BOX products).
    Bundle SKUs are skipped because they won't match the single-item
    VLOOKUP pattern.

    Column B (ITEM) is written in the same type (int vs str) as the existing
    Data1 rows so that downstream VLOOKUPs into Data2 remain type-consistent.
    """
    ws = replenish_wb["Data1"]

    item_col_type = _detect_col_type(ws, column=2)

    existing_skus: set[str] = set()
    for excel_row in range(2, ws.max_row + 1):
        val = ws.cell(row=excel_row, column=1).value
        if val:
            existing_skus.add(str(val).strip())

    next_row = ws.max_row + 1
    added = 0
    added_skus: list[str] = []

    for r in rows:
        if r.get("is_bundle"):
            continue
        sku_no_prefix = r["sku_no_prefix"]
        if sku_no_prefix in existing_skus:
            continue

        item_no = extract_item_number(r["sku"])
        uom_val = r.get("wh_uom", r.get("uom", 1))

        item_val = str(item_no) if item_col_type == "str" else item_no

        ws.cell(row=next_row, column=1, value=sku_no_prefix)       # A  SKU
        ws.cell(row=next_row, column=2, value=item_val)             # B  ITEM
        ws.cell(row=next_row, column=3, value=uom_val)              # C  UOM (numeric)

        existing_skus.add(sku_no_prefix)
        added_skus.append(sku_no_prefix)
        next_row += 1
        added += 1

    if added:
        print(f"  Data1 sheet: {added} new row(s) added (ITEM col type: {item_col_type})")
        for s in added_skus:
            print(f"    + {s}")


def update_data2_sheet(replenish_wb, rows, item_df):
    """Append rows to the Data2 sheet for any Item numbers not already present,
    and overwrite existing rows whose Case Qty / Box Qty cells contain external
    VLOOKUP formulas or are empty (these would fail when the workbook is
    opened standalone and cause columns N and O in *For WH* to stay blank).

    Data2 layout: A=Item No., B=Item Description, C=Case Qty, D=Box Qty.
    Case/Box quantities come from the Item List (with CASE_QTY_OVERRIDES).
    Item Description is left blank for auto-added rows.

    Column A (Item No.) is ALWAYS written as int so that the plain VLOOKUP
    in 'For WH' columns N and O can reliably match — mixing text and numeric
    item numbers in column A causes VLOOKUP to miss rows silently.
    """
    ws = replenish_wb["Data2"]

    # --- Normalize ALL existing Data2 column A values to int ---
    # Some rows may have been stored as text (e.g. "582") while Data1 returns
    # a numeric item number into column D of 'For WH'.  A plain VLOOKUP is
    # type-strict, so text "582" != number 582.  Overwrite every text-string
    # cell in column A with its integer equivalent so the entire column is
    # consistently numeric.
    normalized_count = 0
    for excel_row in range(2, ws.max_row + 1):
        val = ws.cell(row=excel_row, column=1).value
        if isinstance(val, str) and val.strip():
            try:
                ws.cell(row=excel_row, column=1, value=int(float(val.strip())))
                normalized_count += 1
            except (TypeError, ValueError):
                pass
    if normalized_count:
        print(f"  Data2 sheet: normalized {normalized_count} text item number(s) in col A to int")

    # Build map of item_no -> excel_row for all existing rows
    existing_item_row: dict[int, int] = {}
    for excel_row in range(2, ws.max_row + 1):
        val = ws.cell(row=excel_row, column=1).value
        if val is not None:
            try:
                existing_item_row[int(float(val))] = excel_row
            except (TypeError, ValueError):
                pass

    all_item_nos: set[int] = set()
    for r in rows:
        if r.get("is_bundle"):
            for cd in r.get("component_details", []):
                if cd.get("item_no") is not None:
                    all_item_nos.add(int(cd["item_no"]))
        else:
            item_no = extract_item_number(r["sku"])
            if item_no is not None:
                all_item_nos.add(item_no)

    next_row = ws.max_row + 1
    added = 0
    updated = 0

    for item_no in sorted(all_item_nos):
        case_qty = CASE_QTY_OVERRIDES.get(item_no)
        box_qty = None

        if case_qty is None:
            match = item_df[item_df["Item No."] == item_no]
            if not match.empty:
                case_qty = _num(match.iloc[0].get("Case Qty"), 0)
                box_qty = _num(match.iloc[0].get("Box Qty"), 0)
            else:
                case_qty = 0
                box_qty = 0
        else:
            match = item_df[item_df["Item No."] == item_no]
            if not match.empty:
                box_qty = _num(match.iloc[0].get("Box Qty"), 0)
            else:
                box_qty = 0

        case_val = int(case_qty) if case_qty else ""
        box_val = int(box_qty) if box_qty else ""
        item_val = item_no  # always int — column A is normalized to int above

        if item_no in existing_item_row:
            # Row already exists — overwrite C and D only when the current
            # value is a formula string (e.g. external VLOOKUP) or missing.
            er = existing_item_row[item_no]
            existing_c = ws.cell(row=er, column=3).value
            existing_d = ws.cell(row=er, column=4).value
            needs_update = False
            if existing_c is None or (isinstance(existing_c, str) and existing_c.startswith("=")):
                ws.cell(row=er, column=3, value=case_val)
                needs_update = True
            if existing_d is None or (isinstance(existing_d, str) and existing_d.startswith("=")):
                ws.cell(row=er, column=4, value=box_val)
                needs_update = True
            if needs_update:
                updated += 1
        else:
            # Item missing — add a new row
            ws.cell(row=next_row, column=1, value=item_val)                          # A
            ws.cell(row=next_row, column=2, value="")                                # B
            ws.cell(row=next_row, column=3, value=case_val)                          # C
            ws.cell(row=next_row, column=4, value=box_val)                           # D
            next_row += 1
            added += 1

    if added or updated:
        print(f"  Data2 sheet: {added} new row(s) added, {updated} row(s) updated")
        for item_no in sorted(all_item_nos - set(existing_item_row.keys())):
            print(f"    + Item #{item_no}")


# ---------------------------------------------------------------------------
# Worksheet-name helper
# ---------------------------------------------------------------------------


def find_sheet(wb, search_term):
    """Return the first worksheet whose name contains *search_term* (case-insensitive)."""
    for name in wb.sheetnames:
        if search_term.lower() in name.lower():
            return wb[name]
    raise KeyError(
        f"No sheet matching '{search_term}' in workbook.  "
        f"Available sheets: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# Output 1 — Amazon FBA Manifest
# ---------------------------------------------------------------------------


MANIFEST_HEADERS = [
    "Merchant SKU",       # A
    "Quantity",           # B
    "Prep owner",         # C  (values left blank)
    "Labeling owner",     # D  (values left blank)
    "Units per box",      # E
    "Number of boxes",    # F
    "Box length (in)",    # G
    "Box width (in)",     # H
    "Box height (in)",    # I
    "Box weight (lb)",    # J
    "Pack Unit",          # K
]


def validate_manifest_num_boxes(rows):
    """Check that Number of boxes = Sellable Unit / Units per box for each row.

    Logs mismatches to console. For CASE: num_boxes should equal Case to pull.
    For BOX: num_boxes should equal Box to pull.
    """
    mismatches = []
    for r in rows:
        sellable = r.get("adj_sets", 0)
        upb = r.get("manifest_units_per_box") or r.get("sets_per_pack")
        num_boxes = r.get("manifest_num_boxes") or r.get("num_packs")
        if upb is None or upb <= 0:
            continue
        expected = sellable / upb
        if abs((num_boxes or 0) - expected) > 0.01:
            mismatches.append(
                f"SKU {r.get('sku', '?')}: Number of boxes mismatch — "
                f"has {num_boxes}, expected {expected:.2f}"
            )
    if mismatches:
        print("\n  FBA Manifest Number of boxes validation:")
        for msg in mismatches:
            print(f"    ⚠️ {msg}")


def write_manifest(manifest_wb, rows, output_path):
    """Create a clean manifest workbook using the Amazon Send-to-Amazon headers.

    Column headers match the official ManifestFileUpload template exactly.
    Prep owner (C) and Labeling owner (D) are left blank per Amazon defaults.
    Pack Unit (K) is appended so the warehouse knows CASE vs BOX.
    """
    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "FBA Manifest"

    # Row 1: headers
    for col, header in enumerate(MANIFEST_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD

    # Row 2+: data
    for idx, r in enumerate(rows):
        rn = 2 + idx
        pack_type = r.get("pack_type", "")
        pack_qty = _num(r.get("pack_qty"), 0)
        uom = _num(r.get("uom", r.get("wh_uom", 1)), 1)
        if uom <= 0:
            uom = 1
        if r.get("is_bundle"):
            units_per_box = r.get("manifest_units_per_box", r.get("sets_per_pack"))
            num_boxes = r.get("manifest_num_boxes", r.get("num_packs"))
        else:
            units_per_box = math.floor(pack_qty / uom) if pack_qty and uom else r.get("sets_per_pack", 1)
            if units_per_box <= 0:
                units_per_box = 1
            num_boxes = r.get("num_packs", r.get("manifest_num_boxes", 1))
        ws.cell(row=rn, column=1, value=r["sku"])          # A  Merchant SKU
        ws.cell(row=rn, column=2, value=r["adj_sets"])     # B  Quantity (sellable units)
        # C  Prep owner — blank
        # D  Labeling owner — blank
        ws.cell(row=rn, column=5, value=units_per_box)     # E  Units per box (per SKU-type rules)
        ws.cell(row=rn, column=6, value=num_boxes)         # F  Number of boxes
        ws.cell(row=rn, column=7, value=r["length"])        # G  Box length (in)
        ws.cell(row=rn, column=8, value=r["width"])         # H  Box width (in)
        ws.cell(row=rn, column=9, value=r["height"])        # I  Box height (in)
        ws.cell(row=rn, column=10, value=r["weight"])       # J  Box weight (lb)
        ws.cell(row=rn, column=11, value=r["pack_type"])    # K  Pack Unit

        fill = FILL_LIGHT if idx % 2 == 0 else FILL_WHITE
        for col in range(1, 12):
            ws.cell(row=rn, column=col).fill = fill

    # Auto-fit column widths for readability
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 12):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 12)

    out_wb.save(output_path)
    print(f"  Manifest saved  -> {output_path}")


# ---------------------------------------------------------------------------
# Output 2 — Warehouse Replenishment Sheet
# ---------------------------------------------------------------------------


def write_replenishment(replenish_wb, rows, output_path):
    """Fill the *For WH* sheet with data + formulas and save as a new file.

    Hard-coded columns: A (SKU), B (ASIN), C (FNSKU), P (Rec Qty), R (Pack Unit).
    All other columns are live Excel formulas that calculate when the file is
    opened in Excel.
    """
    ws = replenish_wb["For WH"]

    # Extra column headers
    ws.cell(row=1, column=17, value="Actual Qty Replenished").font = BOLD  # Q
    ws.cell(row=1, column=18, value="Current WH Inventory").font = BOLD    # R
    ws.cell(row=1, column=19, value="Pack Unit").font = BOLD               # S

    last_data_row = 1 + len(rows)

    for idx, r in enumerate(rows):
        rn = 2 + idx          # Excel row number
        rn_s = str(rn)        # stringified for formula interpolation

        # --- Hard-coded input columns ---
        ws.cell(row=rn, column=1, value=r["sku_no_prefix"])        # A  SKU
        ws.cell(row=rn, column=2, value=r["asin"])                 # B  ASIN
        ws.cell(row=rn, column=3, value=r["fnsku"])                # C  FNSKU
        ws.cell(row=rn, column=16, value=r["inv_to_send"])        # P  Rec Replenishment Qty
        ws.cell(row=rn, column=17, value=r["adj_sets"])            # Q  Actual Qty Replenished
        ws.cell(row=rn, column=19, value=r["pack_type"])           # S  Pack Unit

        if r.get("is_bundle"):
            # --- Bundle row: hardcode columns that can't use VLOOKUP ---
            item_nums = ", ".join(
                str(cd["item_no"]) for cd in r["component_details"]
            )
            ws.cell(row=rn, column=4, value=item_nums)                 # D  Item #
            ws.cell(row=rn, column=5, value=r["total_ea"])             # E  Total EA (bundles)
            ws.cell(row=rn, column=6, value=r["adj_sets"])             # F  Sellable (bundles)

            box_packs = sum(
                cd["comp_packs"] for cd in r["component_details"]
                if cd["pack_type"] == "BOX"
            )
            case_packs = sum(
                cd["comp_packs"] for cd in r["component_details"]
                if cd["pack_type"] == "CASE"
            )
            # Mutually exclusive: CASE → only Case to pull; BOX → only Box to pull
            pack_type = r.get("pack_type", "")
            ws.cell(row=rn, column=7, value=(box_packs or "") if pack_type == "BOX" else "")   # G  Box to pull
            ws.cell(row=rn, column=8, value=(case_packs or "") if pack_type == "CASE" else "") # H  Case to pull
            ws.cell(row=rn, column=9, value=r.get("sets_per_pack", ""))  # I  Sets (bundles)
            ws.cell(row=rn, column=10, value="")                      # J  Label
            ws.cell(row=rn, column=11, value="")                      # K  Packing Instructions
            ws.cell(row=rn, column=12, value="")                      # L  Supplies
            ws.cell(row=rn, column=13, value=r["bundle_uom"])         # M  UoM
        else:
            # --- Regular row: E, F, I as Excel formulas (V5); G, H from formulas ---
            ws.cell(row=rn, column=4,                                  # D  Item #
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Data1!A:B,2,0),"")')

            ws.cell(row=rn, column=5,                                  # E  Total EA = M*Q
                    value=f"=M{rn_s}*Q{rn_s}")

            ws.cell(row=rn, column=6,                                  # F  Sellable Unit = I*(H or G)
                    value=f'=I{rn_s}*IF(S{rn_s}="CASE",H{rn_s},G{rn_s})')

            # Mutually exclusive: Pack Unit=CASE → only Case to pull; Pack Unit=BOX → only Box to pull
            ws.cell(row=rn, column=7,                                  # G  Box to pull
                    value=f'=IF(S{rn_s}="CASE","",IFERROR(ROUNDUP(E{rn_s}/N{rn_s},0),""))')

            ws.cell(row=rn, column=8,                                  # H  Case to pull
                    value=f'=IF(S{rn_s}="BOX","",IFERROR(ROUNDUP(E{rn_s}/O{rn_s},0),""))')

            ws.cell(row=rn, column=9,                                  # I  Sets per Box/Case
                    value=f'=IF(S{rn_s}="CASE",FLOOR(O{rn_s}/M{rn_s},1),FLOOR(N{rn_s}/M{rn_s},1))')

            ws.cell(row=rn, column=10,                                 # J  Label
                    value=f'=IF(ISNUMBER(SEARCH("FNSKU",C{rn_s})),"Labeling","")')

            fb = r.get("instruction_fallback")
            if fb and fb.get("instruction_fba"):
                ws.cell(row=rn, column=11, value=fb["instruction_fba"])
            else:
                ws.cell(row=rn, column=11,                             # K  Packing Instructions
                        value=f'=IFERROR(VLOOKUP(A{rn_s},Instruction!A:F,6,0),"")')

            ws.cell(row=rn, column=12,                                 # L  Supplies
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Instruction!A:J,10,0),"No Supplies")')

            ws.cell(row=rn, column=13,                                 # M  UoM
                    value=f'=IFERROR(VLOOKUP(A{rn_s},Data1!A:C,3,0),1)')

        # N and O: VLOOKUP from Data2; applies to all rows including bundles.
        # Using VLOOKUP instead of INDEX/MATCH+TEXT(range) avoids needing
        # array-formula entry (Ctrl+Shift+Enter), which openpyxl cannot write
        # and which silently returns "" on pre-365 Excel.
        ws.cell(row=rn, column=14,                                     # N  BOX
                value=f'=IFERROR(VLOOKUP(D{rn_s},Data2!$A:$D,4,0),"")')
        ws.cell(row=rn, column=15,                                     # O  CASE
                value=f'=IFERROR(VLOOKUP(D{rn_s},Data2!$A:$C,3,0),"")')

        ws.cell(row=rn, column=18, value="")                           # R  Current WH Inventory

        # Alternating row fill + thin borders
        fill = FILL_LIGHT if idx % 2 == 0 else FILL_WHITE
        for col in range(1, 20):
            cell = ws.cell(row=rn, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER

    # --- Totals row ---
    tr = last_data_row + 1
    ws.cell(row=tr, column=1, value="TOTALS").font = BOLD
    for col_idx in (5, 6, 7, 8):  # E, F, G, H
        letter = chr(ord("A") + col_idx - 1)
        cell = ws.cell(row=tr, column=col_idx,
                       value=f"=SUM({letter}2:{letter}{last_data_row})")
        cell.font = BOLD

    # --- Update Table1 reference to cover header + data + totals ---
    try:
        table = ws.tables["Table1"]
        table.ref = f"A1:S{tr}"
    except (KeyError, AttributeError):
        for tname in list(ws.tables):
            ws.tables[tname].ref = f"A1:S{tr}"
            break

    replenish_wb.save(output_path)
    print(f"  Replenishment saved -> {output_path}")


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------


def print_summary(rows, missing_instructions, missing_pack_data):
    """Print a human-readable processing summary to stdout."""
    total_items = len(rows)
    total_ea = sum(r["total_ea"] for r in rows)
    cases = [r for r in rows if r["pack_type"] == "CASE"]
    boxes = [r for r in rows if r["pack_type"] == "BOX"]
    bundles = [r for r in rows if r.get("is_bundle")]

    print(f"\n{'=' * 58}")
    print(f"  Total items processed  : {total_items}")
    print(f"  Total EA               : {total_ea:,}")
    print(f"  CASE decisions         : {len(cases)}")
    print(f"  BOX  decisions         : {len(boxes)}")
    print(f"  Bundles                : {len(bundles)}")

    if bundles:
        print("  Bundle details:")
        for r in bundles:
            comps = ", ".join(
                f"{cd['item_no']}×{cd['comp_packs']}{cd['pack_type'][0]}"
                for cd in r["component_details"]
            )
            print(f"    - {r['sku']}  adj_sets={r['adj_sets']}  "
                  f"sellable_set={r['sellable_set']}  [{comps}]")

    if boxes:
        print("  Items shipped as BOX:")
        for r in boxes:
            if not r.get("is_bundle"):
                print(f"    - {r['sku']}  ({r['total_ea']} EA, "
                      f"{r['num_packs']} box(es))")

    print(f"{'=' * 58}")

    if missing_instructions:
        print("\n  Instruction fallbacks used:")
        for m in missing_instructions:
            fba_val = m.get("instruction_fba", "N/A")
            print(f"    {m['sku']}  -> suffix '{m['suffix_used']}'  "
                  f"-> \"{fba_val}\"")

    if missing_pack_data:
        print("\n  WARNING — no pack-unit data found in Item List for:")
        for sku in missing_pack_data:
            print(f"    - {sku}")

    remainder_warnings = [r for r in rows if r.get("remainder_flag")]
    if remainder_warnings:
        print("\n  Remainder warnings (case/box not a clean multiple of UoM):")
        for r in remainder_warnings:
            print(f"    - {r['sku']}: {r['remainder_flag']}")

    rounding_flags = [r for r in rows if r.get("rounding_flag")]
    if rounding_flags:
        print("\n  Bundle rounding warnings (gap > 20%):")
        for r in rounding_flags:
            print(f"    - {r['rounding_flag']}")

    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    """Entry point — orchestrates the full replenishment pipeline."""
    today = date.today().isoformat()

    print("=" * 58)
    print("  FBA Replenishment Automation")
    print(f"  Date : {today}")
    print(f"  Dir  : {os.path.abspath(CONFIG['input_dir'])}")
    print("=" * 58)

    # ---- 1. Load source files ----
    print("\nLocating source files ...")
    sources = load_sources()

    bts_df = sources["bts_df"]
    item_df = sources["item_df"]
    data1_df = sources["data1_df"]
    instruction_df = sources["instruction_df"]
    bundles_df = sources["bundles_df"]

    # ---- 2. Filter rows with inventory to send ----
    inv_col = "Inv to Send from Warehouse"
    bts_df = bts_df[bts_df[inv_col] > 0].copy()
    if CONFIG["sample_size"]:
        bts_df = bts_df.head(CONFIG["sample_size"])

    print(f"\n  Rows with inventory > 0 : {len(bts_df)}")

    # ---- 3. Deduplicate BTS rows ----
    bts_dupes = bts_df.duplicated(subset=["Merchant SKU"], keep="first")
    if bts_dupes.any():
        n = bts_dupes.sum()
        print(f"  NOTE: dropped {n} duplicate Merchant SKU rows from BTS Calcs")
    bts_df = bts_df.drop_duplicates(subset=["Merchant SKU"], keep="first")

    # ---- 4. Split into bundles vs regular items ----
    bundle_mask = bts_df["Merchant SKU"].str.contains("+", na=False, regex=False)
    bundle_bts = bts_df[bundle_mask].copy()
    regular_bts = bts_df[~bundle_mask].copy()
    print(f"  Regular items           : {len(regular_bts)}")
    print(f"  Bundle items            : {len(bundle_bts)}")

    # ---- 5. Prepare Item List ----
    item_df["Item No."] = pd.to_numeric(item_df["Item No."], errors="coerce")
    item_df = item_df.dropna(subset=["Item No."])
    item_df["Item No."] = item_df["Item No."].astype(int)
    item_df = item_df.drop_duplicates(subset=["Item No."], keep="first")

    # ---- 6. Build UoM lookup from Data1 (col 0 = SKU, col 2 = UOM) ----
    sku_to_uom = dict(
        zip(
            data1_df.iloc[:, 0].astype(str),
            data1_df.iloc[:, 2].astype(str),
        )
    )

    # ---- 7. Set of SKUs present in the Instruction sheet ----
    instruction_skus = set(
        instruction_df.iloc[:, 0].astype(str).str.strip()
    )

    rows: list[dict] = []
    missing_instructions: list[dict] = []
    missing_pack_data: list[str] = []

    # ---- 8. Process regular items first ----
    regular_bts["Item No."] = regular_bts["Merchant SKU"].apply(extract_item_number)
    regular_bts = regular_bts.dropna(subset=["Item No."])
    regular_bts["Item No."] = regular_bts["Item No."].astype(int)

    merged = regular_bts.merge(item_df, on="Item No.", how="left")

    rtype = CONFIG["replenish_type"]
    instr_col_name = "Instruction FBA" if rtype == "FBA" else "Instruction WFS"

    for _, row in merged.iterrows():
        sku = str(row["Merchant SKU"])
        sku_no_prefix = sku.replace("FBA_", "")
        uom_str = sku_to_uom.get(sku_no_prefix, "EACH")
        uom = parse_uom(uom_str)

        has_case_qty = pd.notna(row.get("Case Qty"))
        has_dims = pd.notna(row.get("Case Length"))
        if not has_case_qty:
            label = "(has dimensions but no Case Qty)" if has_dims else "(not in Item List)"
            missing_pack_data.append(f"{sku}  {label}")

        case_dims = {
            "length": row.get("Case Length", 0),
            "width": row.get("Case Width", 0),
            "height": row.get("Case Height", 0),
            "weight": row.get("Case Weight", 0),
        }

        box_dims = (
            {
                "length": row.get("Box Length", 0),
                "width": row.get("Box Width", 0),
                "height": row.get("Box Height", 0),
                "weight": row.get("Box Weight", 0),
            }
            if pd.notna(row.get("Box Qty"))
            else None
        )

        item_no = int(row["Item No."])
        case_qty_val = CASE_QTY_OVERRIDES.get(item_no, row.get("Case Qty", 1))
        box_qty_val = row.get("Box Qty")

        suffix_upper = extract_suffix(sku).upper()
        if suffix_upper == "CASE":
            if item_no in CASE_QTY_OVERRIDES:
                # 8004/8005/8006: treat as singular (each), 5 per case, WH UoM = 1
                uom = 1
                result = select_pack_unit(
                    inv_to_send=row[inv_col],
                    case_qty=CASE_QTY_OVERRIDES[item_no],
                    box_qty=None,
                    uom=uom,
                    case_dims=case_dims,
                    box_dims=box_dims,
                )
                result["wh_uom"] = 1
                result["manifest_units_per_box"] = CASE_QTY_OVERRIDES[item_no]
                result["manifest_num_boxes"] = result["adj_sets"] // CASE_QTY_OVERRIDES[item_no]
            else:
                # Other CASE: no rounding, ship exactly inv_to_send cases; WH UoM = case_qty
                result = _build_case_only_result(
                    inv_to_send=row[inv_col],
                    case_qty=case_qty_val,
                    case_dims=case_dims,
                )
        elif suffix_upper == "BOX":
            # BOX: round to boxes_per_case, always ship cases; WH UoM = box_qty
            result = _build_box_suffix_result(
                inv_to_send=row[inv_col],
                case_qty=case_qty_val,
                box_qty=box_qty_val,
                case_dims=case_dims,
            )
        else:
            # EACH / SET{N} / PACK{N}: standard LCM rounding
            uom = parse_uom(sku_to_uom.get(sku_no_prefix, "EACH"))
            result = select_pack_unit(
                inv_to_send=row[inv_col],
                case_qty=case_qty_val,
                box_qty=box_qty_val,
                uom=uom,
                case_dims=case_dims,
                box_dims=box_dims,
            )

        result["sku"] = sku
        result["sku_no_prefix"] = sku_no_prefix
        result["inv_to_send"] = int(_num(row[inv_col]))
        result["uom"] = result.get("wh_uom", result.get("uom", 1))
        result["is_bundle"] = False
        result["asin"] = (
            str(row["ASIN"]) if pd.notna(row.get("ASIN")) else ""
        )
        result["fnsku"] = (
            str(row["FNSKU"]) if pd.notna(row.get("FNSKU")) else ""
        )

        needs_fallback = False
        if sku_no_prefix not in instruction_skus:
            needs_fallback = True
        else:
            match = instruction_df[instruction_df.iloc[:, 0].astype(str).str.strip() == sku_no_prefix]
            if not match.empty:
                val = match.iloc[0].get(instr_col_name, "")
                if pd.isna(val) or val == "" or val == 0:
                    needs_fallback = True

        if needs_fallback:
            fallback = get_instruction_fallback(sku, instruction_df)
            if fallback:
                result["instruction_fallback"] = fallback
                missing_instructions.append({"sku": sku, **fallback})
            else:
                result["instruction_fallback"] = None
                print(f"  WARNING: no instruction match for {sku}")
        else:
            result["instruction_fallback"] = None

        rows.append(result)

    # ---- 9. Process bundle items (at end) ----
    for _, brow in bundle_bts.iterrows():
        sku = str(brow["Merchant SKU"])
        sku_no_prefix = sku.replace("FBA_", "")

        result = process_bundle(sku, brow[inv_col], item_df, bundles_df)
        result["sku"] = sku
        result["sku_no_prefix"] = sku_no_prefix
        result["inv_to_send"] = int(_num(brow[inv_col]))
        result["asin"] = (
            str(brow["ASIN"]) if pd.notna(brow.get("ASIN")) else ""
        )
        result["fnsku"] = (
            str(brow["FNSKU"]) if pd.notna(brow.get("FNSKU")) else ""
        )
        result["instruction_fallback"] = None
        result["uom"] = result.get("bundle_uom", 1)

        print(f"    BUNDLE {sku_no_prefix}: sellable_set={result['sellable_set']}, "
              f"inv={brow[inv_col]} -> adj={result['adj_sets']}, "
              f"packs={result['num_packs']}")
        rows.append(result)

    if not rows:
        print("\nNo items to process — exiting.")
        return

    # ---- 7b. Round Actual Qty to unit multiple (Fix 4: 17090-A-EACH etc.) ----
    unit_rounding_flags = []
    for r in rows:
        if r.get("is_bundle"):
            continue
        actual = r.get("adj_sets", 0)
        pack_type = r.get("pack_type", "")
        pack_qty = r.get("pack_qty", 0)
        uom = r.get("uom", 1) or 1
        unit_qty = math.floor(pack_qty / uom) if uom and pack_qty else 0
        if unit_qty <= 0:
            unit_qty = 1
        rounded, flag = round_to_unit_multiple(actual, unit_qty, r.get("sku", "?"))
        if rounded != actual:
            r["adj_sets"] = int(rounded)
            r["total_ea"] = int(rounded * uom)
            upb = r.get("manifest_units_per_box") or r.get("sets_per_pack")
            if upb and upb > 0:
                r["manifest_num_boxes"] = math.ceil(rounded / upb)
            r["num_packs"] = r.get("manifest_num_boxes", r.get("num_packs", 1))
        if flag:
            unit_rounding_flags.append(flag)
    if unit_rounding_flags:
        print("\n  Actual Qty rounding (below minimum pack):")
        for f in unit_rounding_flags:
            print(f"    {f}")

    # ---- 8. Write output files ----
    manifest_path = unique_path(os.path.join(
        CONFIG["output_dir"],
        CONFIG["manifest_output"].format(date=today),
    ))
    replenish_path = unique_path(os.path.join(
        CONFIG["output_dir"],
        CONFIG["replenish_output"].format(date=today),
    ))

    # ---- 8a. Update Instruction sheet with fallback values ----
    update_instruction_sheet(
        sources["replenish_wb"], rows, instruction_df, CONFIG["replenish_type"]
    )

    # ---- 8b. Auto-populate Data1 / Data2 with missing entries ----
    update_data1_sheet(sources["replenish_wb"], rows)
    update_data2_sheet(sources["replenish_wb"], rows, item_df)

    print("\nWriting output files ...")
    validate_manifest_num_boxes(rows)
    write_manifest(sources["manifest_wb"], rows, manifest_path)
    write_replenishment(sources["replenish_wb"], rows, replenish_path)

    # ---- 9. Summary ----
    print_summary(rows, missing_instructions, missing_pack_data)

    print(f"  Manifest  -> {os.path.abspath(manifest_path)}")
    print(f"  Replenish -> {os.path.abspath(replenish_path)}")
    print("\nDone.")


if __name__ == "__main__":
    main()
